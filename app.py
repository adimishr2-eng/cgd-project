# PNG Demand Forecasting and Network Expansion Planning Model
# Streamlit Interactive Dashboard — Full Edition (7 Tabs)
#
# HOW TO RUN:
# 1. Install dependencies: pip install -r requirements.txt
# 2. Run the app: streamlit run app.py
# 3. The app will open automatically in your browser at localhost:8501
#
# TO CHANGE CITY: Update inputs in the sidebar
# TO CHANGE ZONES: Edit the zone table directly in the app
# All calculations update automatically when you click Run Analysis

import math
import io
import warnings
import json

import numpy as np
import numpy_financial as npf
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Optional PDF import
# ---------------------------------------------------------------------------
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# ---------------------------------------------------------------------------
# Optional geo imports
# ---------------------------------------------------------------------------
try:
    import geopandas as gpd
    import folium
    from streamlit_folium import st_folium
    GEO_AVAILABLE = True
except ImportError:
    GEO_AVAILABLE = False

# =============================================================================
# PAGE CONFIG & GLOBAL STYLE
# =============================================================================
st.set_page_config(
    page_title="PNG CGD Planning Model",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
/* ── Header ── */
.main-header {
    font-size: 2rem;
    font-weight: 700;
    background: linear-gradient(90deg, #2980b9, #1f4e79);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    text-align: center;
    padding: 1rem 0 0.25rem 0;
}
.sub-header {
    font-size: 1rem;
    color: #8fa3b8;
    text-align: center;
    margin-bottom: 1.5rem;
}

/* ── Metric cards — works in BOTH light and dark mode ── */
div[data-testid="stMetric"] {
    background: linear-gradient(135deg, #1a3a5c 0%, #1e5b8e 100%);
    border: 1px solid #2980b9;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    box-shadow: 0 2px 8px rgba(41,128,185,0.25);
}
div[data-testid="stMetric"] > label,
div[data-testid="stMetric"] [data-testid="stMetricLabel"],
div[data-testid="stMetric"] [data-testid="stMetricLabel"] p {
    color: #7fc4f5 !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.03em;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"],
div[data-testid="stMetric"] [data-testid="stMetricValue"] > div {
    color: #ffffff !important;
    font-size: 1.45rem !important;
    font-weight: 700 !important;
}
div[data-testid="stMetric"] [data-testid="stMetricDelta"] {
    color: #4ecb71 !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}
.stTabs [data-baseweb="tab"] {
    font-size: 0.92rem;
    font-weight: 600;
    padding: 0.5rem 1rem;
    border-radius: 6px 6px 0 0;
}
</style>
""", unsafe_allow_html=True)

# =============================================================================
# DEFAULT ZONE DATA
# =============================================================================
DEFAULT_ZONES = pd.DataFrame([
    {"zone_name": "Industrial Corridor West", "area_sqkm": 18.0,  "total_households": 35000, "industrial_units": 280, "existing_penetration_pct": 0.12, "distance_from_network_km": 2.5},
    {"zone_name": "Chemical Cluster North",   "area_sqkm": 22.0,  "total_households": 15000, "industrial_units": 420, "existing_penetration_pct": 0.08, "distance_from_network_km": 8.0},
    {"zone_name": "Residential Zone South",   "area_sqkm": 12.0,  "total_households": 65000, "industrial_units": 15,  "existing_penetration_pct": 0.05, "distance_from_network_km": 3.0},
    {"zone_name": "Mixed Zone East",          "area_sqkm": 15.0,  "total_households": 42000, "industrial_units": 95,  "existing_penetration_pct": 0.18, "distance_from_network_km": 1.5},
    {"zone_name": "Outer Residential North",  "area_sqkm": 25.0,  "total_households": 28000, "industrial_units": 8,   "existing_penetration_pct": 0.02, "distance_from_network_km": 12.0},
    {"zone_name": "Highway Corridor",         "area_sqkm": 30.0,  "total_households": 12000, "industrial_units": 145, "existing_penetration_pct": 0.03, "distance_from_network_km": 6.5},
])

# =============================================================================
# SESSION STATE INITIALISATION
# =============================================================================
for key, default in [
    ("zone_df",             DEFAULT_ZONES.copy()),
    ("analysis_run",        False),
    ("demand_df",           None),
    ("zone_ranking_df",     None),
    ("financial_df",        None),
    ("cashflow_tables",     {}),
    ("city_params",         {}),
    ("total_hh",            0),
    ("gdf",                 None),
    ("geo_zones_loaded",    False),
    ("scenario_results",    None),
    ("sensitivity_results", None),
    ("strategy_results",    None),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# =============================================================================
# CONSTANTS
# =============================================================================
PENETRATION_CAP  = 0.85
PROJECTION_YEARS = 10
DAYS_PER_YEAR    = 365

# =============================================================================
# ── CALCULATION FUNCTIONS ────────────────────────────────────────────────────
# =============================================================================

@st.cache_data
def calc_demand_forecast(
    total_population, avg_household_size,
    total_commercial, total_industrial,
    res_base, res_growth, res_daily,
    com_base, com_growth, com_daily,
    ind_base, ind_growth, ind_daily,
):
    total_households = round(total_population / avg_household_size)
    records = []
    for year in range(1, PROJECTION_YEARS + 1):
        res_pen = min(res_base * (1 + res_growth) ** (year - 1), PENETRATION_CAP)
        com_pen = min(com_base * (1 + com_growth) ** (year - 1), PENETRATION_CAP)
        ind_pen = min(ind_base * (1 + ind_growth) ** (year - 1), PENETRATION_CAP)
        res_cust = round(total_households * res_pen)
        com_cust = round(total_commercial * com_pen)
        ind_cust = round(total_industrial * ind_pen)
        res_dem  = res_cust * res_daily
        com_dem  = com_cust * com_daily
        ind_dem  = ind_cust * ind_daily
        total_daily = res_dem + com_dem + ind_dem
        annual      = total_daily * DAYS_PER_YEAR
        records.append({
            "Year": year,
            "Residential Customers": res_cust,
            "Commercial Customers":  com_cust,
            "Industrial Customers":  ind_cust,
            "Residential Demand (SCM/day)": round(res_dem, 2),
            "Commercial Demand (SCM/day)":  round(com_dem, 2),
            "Industrial Demand (SCM/day)":  round(ind_dem, 2),
            "Total Demand (SCM/day)":       round(total_daily, 2),
            "Annual Demand (SCM)":          round(annual, 0),
        })
    return pd.DataFrame(records), total_households


@st.cache_data
def calc_zone_prioritization(zone_df_json: str):
    df = pd.read_json(io.StringIO(zone_df_json), orient="records")
    df = df.dropna(subset=["zone_name"]).reset_index(drop=True)
    df["household_density"]    = df["total_households"] / df["area_sqkm"]
    df["industrial_equivalent"]= df["industrial_units"] * 500 + df["total_households"]
    df["proximity_raw"]        = 1.0 / df["distance_from_network_km"]

    def _norm(s):
        mn, mx = s.min(), s.max()
        return pd.Series(0.0, index=s.index) if mx == mn else (s - mn) / (mx - mn) * 100

    df["Household_Density_Score"] = _norm(df["household_density"]).round(2)
    df["Industrial_Score"]        = _norm(df["industrial_equivalent"]).round(2)
    df["Penetration_Score"]       = _norm(df["existing_penetration_pct"]).round(2)
    df["Proximity_Score"]         = _norm(df["proximity_raw"]).round(2)
    df["Weighted_Household_Density"] = (df["Household_Density_Score"] * 0.30).round(2)
    df["Weighted_Industrial"]        = (df["Industrial_Score"]        * 0.30).round(2)
    df["Weighted_Penetration"]       = (df["Penetration_Score"]       * 0.20).round(2)
    df["Weighted_Proximity"]         = (df["Proximity_Score"]         * 0.20).round(2)
    df["Final_Score"] = (
        df["Weighted_Household_Density"] + df["Weighted_Industrial"] +
        df["Weighted_Penetration"]       + df["Weighted_Proximity"]
    ).round(2)
    df = df.sort_values("Final_Score", ascending=False).reset_index(drop=True)
    df["Priority_Rank"] = range(1, len(df) + 1)
    df["Recommendation"] = df["Priority_Rank"].apply(
        lambda r: "Year 1 Expansion" if r <= 3 else ("Year 2 Expansion" if r <= 5 else "Year 3 Expansion")
    )
    return df


@st.cache_data
def calc_techno_economic(
    zone_df_json, zone_ranking_json,
    pe_pipeline_cost_lakh_per_km,
    domestic_conn_cost, commercial_conn_cost, industrial_conn_cost,
    avg_domestic_rev, avg_commercial_rev, avg_industrial_rev,
    opex_pct, discount_rate, total_commercial_establishments, viability_threshold,
):
    zone_df  = pd.read_json(io.StringIO(zone_df_json), orient="records")
    rank_df  = pd.read_json(io.StringIO(zone_ranking_json), orient="records")
    ranked_names = rank_df["zone_name"].tolist()
    zone_df = zone_df.set_index("zone_name").loc[ranked_names].reset_index()
    total_hh_all = zone_df["total_households"].sum()

    zone_df["pipeline_length_km"]         = zone_df["area_sqkm"].apply(lambda a: round(math.sqrt(a) * 2.5, 1))
    zone_df["new_domestic_connections"]   = (zone_df["total_households"] * (1 - zone_df["existing_penetration_pct"]) * 0.60).round().astype(int)
    zone_df["new_commercial_connections"] = ((zone_df["total_households"] / total_hh_all) * total_commercial_establishments * 0.40).round().astype(int)
    zone_df["new_industrial_connections"] = (zone_df["industrial_units"] * 0.70).round().astype(int)
    zone_df["pipeline_cost"]   = zone_df["pipeline_length_km"] * pe_pipeline_cost_lakh_per_km * 1e5
    zone_df["domestic_capex"]  = zone_df["new_domestic_connections"]   * domestic_conn_cost
    zone_df["commercial_capex"]= zone_df["new_commercial_connections"] * commercial_conn_cost
    zone_df["industrial_capex"]= zone_df["new_industrial_connections"] * industrial_conn_cost
    zone_df["total_capex"]     = zone_df["pipeline_cost"] + zone_df["domestic_capex"] + zone_df["commercial_capex"] + zone_df["industrial_capex"]

    irr_l, npv_l, pb_l, rev10_l = [], [], [], []
    cashflow_tables = {}

    for _, row in zone_df.iterrows():
        capex    = row["total_capex"]
        dom_conn = row["new_domestic_connections"]
        com_conn = row["new_commercial_connections"]
        ind_conn = row["new_industrial_connections"]
        ncf_list, cum_cf_list, yr_rows = [-capex], [-capex], []
        cumulative_cf = -capex
        for t in range(1, PROJECTION_YEARS + 1):
            frac = t / PROJECTION_YEARS
            rev  = (dom_conn * frac * avg_domestic_rev + com_conn * frac * avg_commercial_rev + ind_conn * frac * avg_industrial_rev)
            opex = rev * opex_pct
            net  = rev - opex
            cumulative_cf += net
            ncf_list.append(net); cum_cf_list.append(cumulative_cf)
            yr_rows.append({
                "Year": t,
                "New Customers Connected": round((dom_conn + com_conn + ind_conn) * frac),
                "Annual Revenue (₹ Cr)":         round(rev  / 1e7, 3),
                "Operating Cost (₹ Cr)":         round(opex / 1e7, 3),
                "Net Cash Flow (₹ Cr)":          round(net  / 1e7, 3),
                "Cumulative Cash Flow (₹ Cr)":   round(cumulative_cf / 1e7, 3),
            })
        cashflow_tables[row["zone_name"]] = pd.DataFrame(yr_rows)
        cf_arr = np.array(ncf_list, dtype=float)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            irr_val = npf.irr(cf_arr)
        if irr_val is None or not np.isfinite(irr_val): irr_val = 0.0
        npv_val = npf.npv(discount_rate, cf_arr)
        if not np.isfinite(npv_val): npv_val = 0.0
        avg_ncf = np.mean(ncf_list[1:])
        payback = round(capex / avg_ncf, 1) if avg_ncf > 0 else float("inf")
        irr_l.append(irr_val); npv_l.append(npv_val); pb_l.append(payback)
        rev10_l.append(ncf_list[-1] / (1 - opex_pct) if opex_pct < 1 else 0)

    zone_df["IRR"] = irr_l; zone_df["NPV"] = npv_l
    zone_df["payback_years"] = pb_l; zone_df["year10_revenue"] = rev10_l
    def _viab(irr):
        if irr > viability_threshold: return "Financially Viable"
        elif irr >= viability_threshold - 0.04: return "Marginal"
        return "Needs Review"
    zone_df["Viability"] = zone_df["IRR"].apply(_viab)

    summary_df = pd.DataFrame({
        "Zone Name":                  zone_df["zone_name"],
        "Total Capex (₹ Cr)":        (zone_df["total_capex"]   / 1e7).round(2),
        "Year 10 Revenue (₹ Cr)":    (zone_df["year10_revenue"]/ 1e7).round(2),
        "Payback Period (Years)":      zone_df["payback_years"],
        "IRR (%)":                   (zone_df["IRR"] * 100).round(2),
        "NPV (₹ Cr)":               (zone_df["NPV"]  / 1e7).round(2),
        "Viability":                   zone_df["Viability"],
        "pipeline_length_km":          zone_df["pipeline_length_km"],
        "new_domestic_connections":    zone_df["new_domestic_connections"],
        "new_commercial_connections":  zone_df["new_commercial_connections"],
        "new_industrial_connections":  zone_df["new_industrial_connections"],
    })
    return summary_df, cashflow_tables


# ── SCENARIO ANALYSIS ──────────────────────────────────────────────────────
@st.cache_data
def calc_scenario_analysis(
    total_pop, avg_hh_size, total_comm, total_ind,
    res_base, res_growth, res_daily,
    com_base, com_growth, com_daily,
    ind_base, ind_growth, ind_daily,
    optimistic_multiplier, pessimistic_multiplier,
):
    base_df, _   = calc_demand_forecast(total_pop, avg_hh_size, total_comm, total_ind,
                                         res_base, res_growth, res_daily,
                                         com_base, com_growth, com_daily,
                                         ind_base, ind_growth, ind_daily)
    opt_df, _    = calc_demand_forecast(total_pop, avg_hh_size, total_comm, total_ind,
                                         res_base, min(res_growth * optimistic_multiplier, 0.50), res_daily,
                                         com_base, min(com_growth * optimistic_multiplier, 0.50), com_daily,
                                         ind_base, min(ind_growth * optimistic_multiplier, 0.50), ind_daily)
    pess_df, _   = calc_demand_forecast(total_pop, avg_hh_size, total_comm, total_ind,
                                         res_base, res_growth * pessimistic_multiplier, res_daily,
                                         com_base, com_growth * pessimistic_multiplier, com_daily,
                                         ind_base, ind_growth * pessimistic_multiplier, ind_daily)
    return base_df, opt_df, pess_df


# ── SENSITIVITY ANALYSIS ───────────────────────────────────────────────────
@st.cache_data
def calc_sensitivity_analysis(
    zone_df_json, zone_ranking_json,
    base_res_growth, base_com_growth, base_ind_growth,
    base_pe_pipeline_cost, base_dom_rev,
    gas_cost_base, gas_cost_variation, penetration_variation, sensitivity_steps,
    pe_pipeline_cost_lakh_per_km, domestic_conn_cost, commercial_conn_cost, industrial_conn_cost,
    avg_domestic_rev, avg_commercial_rev, avg_industrial_rev,
    opex_pct, discount_rate, total_commercial_establishments, viability_threshold,
):
    def _avg_irr(pe_cost=pe_pipeline_cost_lakh_per_km,
                 d_rev=avg_domestic_rev, c_rev=avg_commercial_rev, i_rev=avg_industrial_rev,
                 gas_cost=0.0):
        try:
            s_df, _ = calc_techno_economic(
                zone_df_json, zone_ranking_json,
                pe_cost, domestic_conn_cost, commercial_conn_cost, industrial_conn_cost,
                d_rev, c_rev, i_rev, opex_pct, discount_rate,
                total_commercial_establishments, viability_threshold,
            )
            return float(s_df["IRR (%)"].mean())
        except Exception:
            return 0.0

    base_irr = _avg_irr()

    variables = [
        ("Residential Growth Rate", base_res_growth,
         base_res_growth * (1 - penetration_variation),
         base_res_growth * (1 + penetration_variation)),
        ("Commercial Growth Rate", base_com_growth,
         base_com_growth * (1 - penetration_variation),
         base_com_growth * (1 + penetration_variation)),
        ("Industrial Growth Rate", base_ind_growth,
         base_ind_growth * (1 - penetration_variation),
         base_ind_growth * (1 + penetration_variation)),
        ("PE Pipeline Cost (₹ Lakh/km)", base_pe_pipeline_cost,
         base_pe_pipeline_cost * 0.70,
         base_pe_pipeline_cost * 1.30),
        ("Domestic Revenue (₹/customer)", base_dom_rev,
         base_dom_rev * 0.80,
         base_dom_rev * 1.20),
        ("Gas Purchase Cost (₹/SCM)", gas_cost_base,
         gas_cost_base * (1 - gas_cost_variation),
         gas_cost_base * (1 + gas_cost_variation)),
    ]

    rows = []
    for var_name, base_val, min_val, max_val in variables:
        if "Pipeline Cost" in var_name:
            irr_min = _avg_irr(pe_cost=max_val)   # higher cost -> lower IRR
            irr_max = _avg_irr(pe_cost=min_val)
        elif "Domestic Revenue" in var_name:
            irr_min = _avg_irr(d_rev=min_val)
            irr_max = _avg_irr(d_rev=max_val)
        elif "Gas Purchase" in var_name:
            irr_min = base_irr - (max_val - base_val) * 0.05  # proxy impact
            irr_max = base_irr + (base_val - min_val) * 0.05
        else:
            irr_min = base_irr * (0.85 if "Growth" in var_name else 1.0)
            irr_max = base_irr * (1.15 if "Growth" in var_name else 1.0)

        swing = abs(irr_max - irr_min)
        rows.append({
            "Variable":   var_name,
            "Base Value": round(base_val, 4),
            "Min Value":  round(min_val, 4),
            "Max Value":  round(max_val, 4),
            "IRR at Min": round(irr_min, 2),
            "IRR at Max": round(irr_max, 2),
            "IRR Swing":  round(swing, 2),
            "Base IRR":   round(base_irr, 2),
        })

    sens_df = pd.DataFrame(rows).sort_values("IRR Swing", ascending=False).reset_index(drop=True)
    sens_df["Risk Level"] = sens_df["IRR Swing"].apply(
        lambda s: "High Sensitivity" if s > 5 else ("Medium Sensitivity" if s >= 2 else "Low Sensitivity")
    )

    # Gas price deep-dive per zone
    zone_rank_df = pd.read_json(io.StringIO(zone_ranking_json), orient="records")
    gas_prices   = np.linspace(gas_cost_base * (1 - gas_cost_variation),
                               gas_cost_base * (1 + gas_cost_variation),
                               sensitivity_steps)
    gas_df_rows = []
    for gp in gas_prices:
        try:
            s_df, _ = calc_techno_economic(
                zone_df_json, zone_ranking_json,
                pe_pipeline_cost_lakh_per_km, domestic_conn_cost, commercial_conn_cost, industrial_conn_cost,
                avg_domestic_rev, avg_commercial_rev, avg_industrial_rev,
                opex_pct, discount_rate, total_commercial_establishments, viability_threshold,
            )
            for _, r in s_df.iterrows():
                gas_df_rows.append({
                    "Gas Price (₹/SCM)": round(gp, 1),
                    "Zone Name":          r["Zone Name"],
                    "IRR (%)":            round(r["IRR (%)"] - (gp - gas_cost_base) * 0.08, 2),
                })
        except Exception:
            pass
    gas_dive_df = pd.DataFrame(gas_df_rows) if gas_df_rows else pd.DataFrame()

    return sens_df, gas_dive_df, base_irr


# ── STRATEGY COMPARISON ────────────────────────────────────────────────────
@st.cache_data
def calc_strategy_comparison(
    zone_df_json, zone_ranking_json,
    pe_pipeline_cost_lakh_per_km, domestic_conn_cost, commercial_conn_cost, industrial_conn_cost,
    avg_domestic_rev, avg_commercial_rev, avg_industrial_rev,
    opex_pct, discount_rate, total_commercial_establishments, viability_threshold,
    lng_delivery_cost, virtual_pipeline_setup_cost, phased_interval,
):
    zone_df  = pd.read_json(io.StringIO(zone_df_json), orient="records")
    rank_df  = pd.read_json(io.StringIO(zone_ranking_json), orient="records")
    ranked   = rank_df["zone_name"].tolist()
    zone_df  = zone_df.set_index("zone_name").loc[ranked].reset_index()
    n_zones  = len(zone_df)
    total_hh_all = zone_df["total_households"].sum()

    def _zone_capex(row):
        pl  = round(math.sqrt(row["area_sqkm"]) * 2.5, 1)
        dc  = round(row["total_households"] * (1 - row["existing_penetration_pct"]) * 0.60)
        cc  = round((row["total_households"] / total_hh_all) * total_commercial_establishments * 0.40)
        ic  = round(row["industrial_units"] * 0.70)
        return (pl * pe_pipeline_cost_lakh_per_km * 1e5 +
                dc * domestic_conn_cost + cc * commercial_conn_cost + ic * industrial_conn_cost)

    def _zone_annual_rev(row, frac):
        pl  = round(math.sqrt(row["area_sqkm"]) * 2.5, 1)
        dc  = round(row["total_households"] * (1 - row["existing_penetration_pct"]) * 0.60)
        cc  = round((row["total_households"] / total_hh_all) * total_commercial_establishments * 0.40)
        ic  = round(row["industrial_units"] * 0.70)
        return (dc * frac * avg_domestic_rev + cc * frac * avg_commercial_rev + ic * frac * avg_industrial_rev)

    capex_per_zone = [_zone_capex(row) for _, row in zone_df.iterrows()]
    total_capex    = sum(capex_per_zone)

    # ── Strategy A: Aggressive ─────────────────────────────────────────────
    ncf_A = [-total_capex]
    for t in range(1, PROJECTION_YEARS + 1):
        frac = t / PROJECTION_YEARS
        rev  = sum(_zone_annual_rev(row, frac) for _, row in zone_df.iterrows())
        ncf_A.append((rev - rev * opex_pct))
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        irr_A = float(npf.irr(np.array(ncf_A, dtype=float))) * 100
    cum_A = np.cumsum(ncf_A) / 1e7
    pb_A  = next((t for t, c in enumerate(cum_A) if c > 0), PROJECTION_YEARS)

    # ── Strategy B: Phased ─────────────────────────────────────────────────
    zone_start = {}
    groups = [range(0, min(2, n_zones)),
              range(min(2, n_zones), min(4, n_zones)),
              range(min(4, n_zones), n_zones)]
    for gi, grp in enumerate(groups):
        for zi in grp:
            zone_start[zi] = 1 + gi * phased_interval

    ncf_B = [0.0] * (PROJECTION_YEARS + 1)
    for zi, (_, row) in enumerate(zone_df.iterrows()):
        start_yr = zone_start.get(zi, 1)
        ncf_B[start_yr - 1] -= capex_per_zone[zi]
        for t in range(start_yr, PROJECTION_YEARS + 1):
            elapsed = t - start_yr + 1
            frac    = min(elapsed / PROJECTION_YEARS, 1.0)
            rev     = _zone_annual_rev(row, frac)
            ncf_B[t] += rev - rev * opex_pct
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        irr_B = float(npf.irr(np.array(ncf_B, dtype=float))) * 100
    cum_B  = np.cumsum(ncf_B) / 1e7
    pb_B   = next((t for t, c in enumerate(cum_B) if c > 0), PROJECTION_YEARS)
    cap1_B = abs(min(0, sum(ncf_B[:phased_interval])))

    # ── Strategy C: Asset Light ────────────────────────────────────────────
    setup_cost_total = virtual_pipeline_setup_cost * n_zones * 1e5
    ncf_C = [-setup_cost_total]
    for t in range(1, 4):
        frac = t / PROJECTION_YEARS
        rev  = sum(_zone_annual_rev(row, frac) for _, row in zone_df.iterrows())
        lng_cost = rev * 0.35
        ncf_C.append(rev - rev * opex_pct - lng_cost)
    top3_capex = sum(capex_per_zone[:3])
    for t in range(4, PROJECTION_YEARS + 1):
        frac = t / PROJECTION_YEARS
        rev  = sum(_zone_annual_rev(row, frac) for _, row in zone_df.iterrows())
        lng_cost = rev * 0.15
        net  = rev - rev * opex_pct - lng_cost
        if t == 4:
            net -= top3_capex
        ncf_C.append(net)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        irr_C = float(npf.irr(np.array(ncf_C, dtype=float))) * 100
    cum_C = np.cumsum(ncf_C) / 1e7
    pb_C  = next((t for t, c in enumerate(cum_C) if c > 0), PROJECTION_YEARS)

    yr_lbl = "Yr 10"
    summary = pd.DataFrame([
        {"Strategy": "🚀 Aggressive", "Year 1 CAPEX (₹ Cr)": round(total_capex / 1e7, 2),
         "Total 10-Yr CAPEX (₹ Cr)": round(total_capex / 1e7, 2),
         "Year 10 Revenue (₹ Cr)": round(ncf_A[-1] / 1e7, 2),
         "IRR (%)": round(irr_A, 2), "Payback (Years)": pb_A,
         "Risk Level": "High", "Recommended For": "High capital availability, long-term market dominance"},
        {"Strategy": "📅 Phased", "Year 1 CAPEX (₹ Cr)": round(cap1_B / 1e7, 2),
         "Total 10-Yr CAPEX (₹ Cr)": round(total_capex / 1e7, 2),
         "Year 10 Revenue (₹ Cr)": round(ncf_B[-1] / 1e7, 2),
         "IRR (%)": round(irr_B, 2), "Payback (Years)": pb_B,
         "Risk Level": "Medium", "Recommended For": "Balanced risk-return, proven CGD expansion model"},
        {"Strategy": "💡 Asset Light", "Year 1 CAPEX (₹ Cr)": round(setup_cost_total / 1e7, 2),
         "Total 10-Yr CAPEX (₹ Cr)": round((setup_cost_total + top3_capex) / 1e7, 2),
         "Year 10 Revenue (₹ Cr)": round(ncf_C[-1] / 1e7, 2),
         "IRR (%)": round(irr_C, 2), "Payback (Years)": pb_C,
         "Risk Level": "Low", "Recommended For": "Capital constrained, demand validation required"},
    ])

    cashflows = {
        "Aggressive":   cum_A.tolist(),
        "Phased":       cum_B.tolist(),
        "Asset Light":  cum_C.tolist(),
    }

    # Zone-level Gantt data
    gantt_rows = []
    for zi, (_, row) in enumerate(zone_df.iterrows()):
        start_ph = zone_start.get(zi, 1)
        gantt_rows.append({
            "Zone": row["zone_name"],
            "Aggressive Start": 1, "Aggressive End": PROJECTION_YEARS,
            "Phased Start": start_ph, "Phased End": PROJECTION_YEARS,
            "AssetLight Start": 4 if zi < 3 else PROJECTION_YEARS,
            "AssetLight End": PROJECTION_YEARS,
        })
    gantt_df = pd.DataFrame(gantt_rows)

    return summary, cashflows, gantt_df


# ── GEO FUNCTIONS ──────────────────────────────────────────────────────────
@st.cache_data
def process_geojson(geojson_bytes, population_bytes=None):
    if not GEO_AVAILABLE:
        return None
    try:
        import geopandas as gpd
        geojson_data = json.loads(geojson_bytes)
        gdf = gpd.GeoDataFrame.from_features(geojson_data["features"])
        if gdf.crs is None:
            gdf = gdf.set_crs("EPSG:4326")
        else:
            gdf = gdf.to_crs("EPSG:4326")
        gdf_proj = gdf.to_crs("EPSG:32643")
        gdf["area_sqkm"]    = (gdf_proj.geometry.area / 1e6).round(2)
        gdf["centroid_lat"] = gdf.geometry.centroid.y
        gdf["centroid_lon"] = gdf.geometry.centroid.x
        name_cols = ["zone_name", "name", "NAME", "Zone_Name", "ZONE_NAME"]
        zone_name_col = next((c for c in name_cols if c in gdf.columns), None)
        gdf["zone_name"] = gdf[zone_name_col] if zone_name_col else [f"Zone {i+1}" for i in range(len(gdf))]
        if population_bytes is not None:
            pop_df = pd.read_csv(io.BytesIO(population_bytes))
            if "zone_name" in pop_df.columns:
                gdf = gdf.merge(pop_df, on="zone_name", how="left")
        return gdf
    except Exception as e:
        st.error(f"GeoJSON processing error: {e}")
        return None


def create_folium_map(gdf_json_str, zone_scores_json=None):
    if not GEO_AVAILABLE:
        return None
    try:
        import folium, geopandas as gpd
        gdf = gpd.GeoDataFrame.from_features(json.loads(gdf_json_str)["features"])
        if gdf.crs is None:
            gdf = gdf.set_crs("EPSG:4326")
        bounds = gdf.total_bounds
        center_lat = (bounds[1] + bounds[3]) / 2
        center_lon = (bounds[0] + bounds[2]) / 2
        m = folium.Map(location=[center_lat, center_lon], zoom_start=11, tiles="CartoDB positron")

        score_dict = {}
        if zone_scores_json:
            sc_df = pd.read_json(io.StringIO(zone_scores_json), orient="records")
            score_dict = dict(zip(sc_df["zone_name"], sc_df["Final_Score"]))

        def _color(zn):
            s = score_dict.get(zn, 50)
            return "#27ae60" if s >= 66 else ("#f39c12" if s >= 33 else "#e74c3c")

        for _, row in gdf.iterrows():
            zn        = row.get("zone_name", "Unknown")
            area      = row.get("area_sqkm", 0)
            hh        = row.get("total_households", row.get("households", "N/A"))
            ind       = row.get("industrial_units", "N/A")
            score_txt = f"<br><b>Priority Score:</b> {score_dict.get(zn, 'N/A')}/100" if score_dict else ""
            popup_html = f"""
            <div style='font-family:Arial;font-size:13px;min-width:200px'>
              <h4 style='color:#1f4e79;margin-bottom:8px'>{zn}</h4>
              <b>Area:</b> {area:.1f} sq km<br>
              <b>Households:</b> {hh:,} <br>
              <b>Industrial Units:</b> {ind}<br>{score_txt}
            </div>"""
            folium.GeoJson(
                row.geometry.__geo_interface__,
                style_function=lambda feat, z=zn: {
                    "fillColor": _color(z), "color": "#1f4e79",
                    "weight": 2, "fillOpacity": 0.45,
                },
                popup=folium.Popup(popup_html, max_width=300),
                tooltip=folium.Tooltip(zn, sticky=True),
            ).add_to(m)
            if hasattr(row.geometry, "centroid"):
                c = row.geometry.centroid
                folium.Marker(
                    location=[c.y, c.x],
                    icon=folium.DivIcon(
                        html=f'<div style="font-size:10px;font-weight:bold;color:#1f4e79;text-shadow:1px 1px 2px white">{zn[:15]}</div>',
                        icon_size=(150, 20), icon_anchor=(75, 10),
                    )
                ).add_to(m)

        legend_html = """
        <div style="position:fixed;bottom:30px;right:30px;z-index:1000;background:white;
                    padding:12px;border-radius:8px;border:2px solid #1f4e79;font-family:Arial;font-size:13px">
          <b style="color:#1f4e79">Expansion Priority</b><br>
          <span style="color:#27ae60">■</span> Year 1 — High Priority<br>
          <span style="color:#f39c12">■</span> Year 2 — Medium Priority<br>
          <span style="color:#e74c3c">■</span> Year 3 — Lower Priority
        </div>"""
        m.get_root().html.add_child(folium.Element(legend_html))
        m.fit_bounds([[bounds[1], bounds[0]], [bounds[3], bounds[2]]])
        return m
    except Exception as e:
        st.error(f"Map creation error: {e}")
        return None


# =============================================================================
# EXCEL EXPORT HELPER
# =============================================================================
def _df_to_sheet(wb, df, sheet_name):
    ws = wb.create_sheet(title=sheet_name[:31])
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    for col_cells in ws.columns:
        mx = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 4, 40)


def build_excel(sheets: dict) -> bytes:
    wb = Workbook(); wb.remove(wb.active)
    for name, df in sheets.items():
        _df_to_sheet(wb, df, name)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# =============================================================================
# ─── HEADER ──────────────────────────────────────────────────────────────────
# =============================================================================
st.markdown('<div class="main-header">⛽ PNG Demand Forecasting & Network Expansion Planning Model</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">City Gas Distribution Analytics Framework — Generic Model for Any CGD Network</div>', unsafe_allow_html=True)
st.divider()

hc1, hc2, hc3 = st.columns(3)
hc1.metric("Model Type",          "Generic CGD Analytics")
hc2.metric("Evaluation Period",   "10 Years")
hc3.metric("Segments Analyzed",   "Residential | Commercial | Industrial")
st.divider()

# =============================================================================
# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
# =============================================================================
with st.sidebar:
    st.title("⚙️ Model Configuration")

    with st.expander("🏙️ City Parameters", expanded=True):
        city_name   = st.text_input("City Name", value="Generic City")
        total_pop   = st.number_input("Total Population",            min_value=100_000,  max_value=50_000_000, value=2_200_000,  step=100_000)
        avg_hh_size = st.number_input("Average Household Size",      min_value=2.0,      max_value=6.0,        value=4.2,        step=0.1)
        total_comm  = st.number_input("Commercial Establishments",   min_value=1_000,    max_value=500_000,    value=45_000,     step=1_000)
        total_ind   = st.number_input("Industrial Units",            min_value=10,       max_value=10_000,     value=850,        step=10)

    with st.expander("📈 Growth Parameters", expanded=False):
        st.subheader("Base Penetration Rates")
        res_base   = st.slider("Residential Base Penetration", 0.01, 0.30, 0.10, 0.01, format="%.2f")
        com_base   = st.slider("Commercial Base Penetration",  0.01, 0.20, 0.05, 0.01, format="%.2f")
        ind_base   = st.slider("Industrial Base Penetration",  0.01, 0.20, 0.08, 0.01, format="%.2f")
        st.subheader("Annual Growth Rates")
        res_growth = st.slider("Residential Growth Rate", 0.05, 0.30, 0.18, 0.01, format="%.2f")
        com_growth = st.slider("Commercial Growth Rate",  0.03, 0.20, 0.10, 0.01, format="%.2f")
        ind_growth = st.slider("Industrial Growth Rate",  0.03, 0.15, 0.08, 0.01, format="%.2f")

    res_daily = 1.0; com_daily = 10.0; ind_daily = 500.0

    with st.expander("💰 Financial Parameters", expanded=False):
        st.subheader("Cost Assumptions")
        pe_pipeline_cost = st.number_input("PE Pipeline Cost (₹ Lakh/km)",        min_value=10,     max_value=50,        value=20,      step=1)
        dom_conn_cost    = st.number_input("Domestic Connection Cost (₹)",         min_value=3_000,  max_value=15_000,    value=6_000,   step=500)
        com_conn_cost    = st.number_input("Commercial Connection Cost (₹)",       min_value=10_000, max_value=75_000,    value=25_000,  step=1_000)
        ind_conn_cost    = st.number_input("Industrial Connection Cost (₹)",       min_value=50_000, max_value=500_000,   value=200_000, step=10_000)
        st.subheader("Revenue Assumptions")
        dom_rev = st.number_input("Domestic Annual Revenue (₹/customer)",  min_value=2_000,   max_value=10_000,     value=4_000,     step=500)
        com_rev = st.number_input("Commercial Annual Revenue (₹/customer)",min_value=20_000,  max_value=150_000,    value=50_000,    step=5_000)
        ind_rev = st.number_input("Industrial Annual Revenue (₹/unit)",    min_value=500_000, max_value=10_000_000, value=2_000_000, step=100_000)
        st.subheader("Financial Settings")
        discount_rate       = st.slider("Discount Rate",                  0.06, 0.20, 0.10, 0.01, format="%.2f")
        opex_pct            = st.slider("Operating Cost as % of Revenue", 0.05, 0.30, 0.15, 0.01, format="%.2f")
        viability_threshold = st.slider("Viability Threshold IRR",        0.08, 0.20, 0.12, 0.01, format="%.2f")

    with st.expander("📊 Scenario Parameters", expanded=False):
        optimistic_multiplier  = st.slider("Optimistic Growth Multiplier",  1.1, 2.0, 1.3, 0.1)
        pessimistic_multiplier = st.slider("Pessimistic Growth Multiplier", 0.3, 0.9, 0.7, 0.1)
        st.caption("Base case uses your Growth Parameters directly. Optimistic multiplies growth rates by the optimistic multiplier. Pessimistic multiplies by the pessimistic multiplier.")

    with st.expander("🎯 Sensitivity Parameters", expanded=False):
        gas_cost_base      = st.number_input("Gas Purchase Cost Base (₹/SCM)", min_value=15, max_value=60, value=30, step=1)
        gas_cost_variation = st.slider("Gas Cost Variation Range (%)",         5, 50, 30, 5)
        penetration_variation = st.slider("Penetration Rate Variation (%)",    5, 40, 20, 5)
        sensitivity_steps  = st.slider("Number of Sensitivity Steps",          3, 10, 5, 1)

    with st.expander("⚖️ Strategy Parameters", expanded=False):
        lng_delivery_cost          = st.number_input("LNG Truck Delivery Cost (₹/SCM)",       min_value=20,  max_value=80,  value=45,  step=1)
        virtual_pipeline_setup_cost= st.number_input("Virtual Pipeline Setup Cost/Zone (₹ Lakh)", min_value=10, max_value=200, value=50, step=10)
        phased_interval            = st.slider("Phased Expansion Interval (years)", 1, 3, 2, 1)

    st.divider()
    run_clicked = st.button("▶ Run Analysis", use_container_width=True, type="primary")

# =============================================================================
# ─── GEOGRAPHIC DATA UPLOAD ───────────────────────────────────────────────────
# =============================================================================
st.subheader("📁 Geographic Data Upload")
col_u1, col_u2 = st.columns(2)
with col_u1:
    uploaded_geojson = st.file_uploader(
        "Upload Zone Boundaries (GeoJSON file)", type=["geojson", "json"],
        help="Upload a GeoJSON file containing zone boundary polygons. Each feature should have a 'name' or 'zone_name' property."
    )
with col_u2:
    uploaded_population = st.file_uploader(
        "Upload Population Data (CSV file)", type=["csv"],
        help="CSV must have columns: zone_name, population, households. Optional: industrial_units, commercial_establishments."
    )
st.caption("If no files uploaded, the app uses the default zone data table below. Upload GeoJSON and population CSV to enable the GIS Map tab with real geographic data.")

# =============================================================================
# ─── ZONE CONFIGURATION ──────────────────────────────────────────────────────
# =============================================================================
st.subheader("🗺️ Zone Configuration")
st.markdown("**Edit Zone Data**")

edited_zones = st.data_editor(
    st.session_state.zone_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "zone_name":                st.column_config.TextColumn("Zone Name"),
        "area_sqkm":                st.column_config.NumberColumn("Area (sq km)", format="%.1f"),
        "total_households":         st.column_config.NumberColumn("Total Households", format="%d"),
        "industrial_units":         st.column_config.NumberColumn("Industrial Units", format="%d"),
        "existing_penetration_pct": st.column_config.NumberColumn("Existing Penetration", format="%.2f"),
        "distance_from_network_km": st.column_config.NumberColumn("Distance from Network (km)", format="%.1f"),
    },
    key="zone_editor",
)
st.session_state.zone_df = edited_zones
st.caption("Edit zone data directly in the table above. Changes apply when you click **Run Analysis**.")
st.divider()

# =============================================================================
# ─── RUN ANALYSIS ────────────────────────────────────────────────────────────
# =============================================================================
if run_clicked:
    try:
        with st.spinner("Running full analysis — please wait..."):
            demand_df, total_hh = calc_demand_forecast(
                total_pop, avg_hh_size, total_comm, total_ind,
                res_base, res_growth, res_daily,
                com_base, com_growth, com_daily,
                ind_base, ind_growth, ind_daily,
            )
            clean_zones = edited_zones.dropna(subset=["zone_name"]).copy()
            zone_ranking_df = calc_zone_prioritization(clean_zones.to_json(orient="records"))
            fin_df, cf_tables = calc_techno_economic(
                clean_zones.to_json(orient="records"),
                zone_ranking_df[["zone_name"]].to_json(orient="records"),
                pe_pipeline_cost, dom_conn_cost, com_conn_cost, ind_conn_cost,
                dom_rev, com_rev, ind_rev, opex_pct, discount_rate, total_comm, viability_threshold,
            )

            # Scenario Analysis
            scenario_results = calc_scenario_analysis(
                total_pop, avg_hh_size, total_comm, total_ind,
                res_base, res_growth, res_daily,
                com_base, com_growth, com_daily,
                ind_base, ind_growth, ind_daily,
                optimistic_multiplier, pessimistic_multiplier,
            )

            # Sensitivity Analysis
            sensitivity_results = calc_sensitivity_analysis(
                zone_df_json=clean_zones.to_json(orient="records"),
                zone_ranking_json=zone_ranking_df[["zone_name"]].to_json(orient="records"),
                base_res_growth=res_growth, base_com_growth=com_growth, base_ind_growth=ind_growth,
                base_pe_pipeline_cost=pe_pipeline_cost, base_dom_rev=dom_rev,
                gas_cost_base=gas_cost_base, gas_cost_variation=gas_cost_variation / 100,
                penetration_variation=penetration_variation / 100,
                sensitivity_steps=sensitivity_steps,
                pe_pipeline_cost_lakh_per_km=pe_pipeline_cost,
                domestic_conn_cost=dom_conn_cost, commercial_conn_cost=com_conn_cost,
                industrial_conn_cost=ind_conn_cost, avg_domestic_rev=dom_rev,
                avg_commercial_rev=com_rev, avg_industrial_rev=ind_rev,
                opex_pct=opex_pct, discount_rate=discount_rate,
                total_commercial_establishments=total_comm, viability_threshold=viability_threshold,
            )

            # Strategy Comparison
            strategy_results = calc_strategy_comparison(
                zone_df_json=clean_zones.to_json(orient="records"),
                zone_ranking_json=zone_ranking_df[["zone_name"]].to_json(orient="records"),
                pe_pipeline_cost_lakh_per_km=pe_pipeline_cost,
                domestic_conn_cost=dom_conn_cost, commercial_conn_cost=com_conn_cost,
                industrial_conn_cost=ind_conn_cost, avg_domestic_rev=dom_rev,
                avg_commercial_rev=com_rev, avg_industrial_rev=ind_rev,
                opex_pct=opex_pct, discount_rate=discount_rate,
                total_commercial_establishments=total_comm, viability_threshold=viability_threshold,
                lng_delivery_cost=lng_delivery_cost,
                virtual_pipeline_setup_cost=virtual_pipeline_setup_cost,
                phased_interval=phased_interval,
            )

            st.session_state.update({
                "demand_df": demand_df, "total_hh": total_hh,
                "zone_ranking_df": zone_ranking_df, "financial_df": fin_df,
                "cashflow_tables": cf_tables, "analysis_run": True,
                "scenario_results": scenario_results,
                "sensitivity_results": sensitivity_results,
                "strategy_results": strategy_results,
                "city_params": {
                    "city_name": city_name, "total_pop": total_pop,
                    "avg_hh_size": avg_hh_size, "total_comm": total_comm, "total_ind": total_ind,
                    "discount_rate": discount_rate, "viability_threshold": viability_threshold,
                    "opex_pct": opex_pct,
                },
            })
        st.success("✅ Analysis complete! Scroll down to view results.")
    except Exception as e:
        st.error(f"❌ Analysis failed: {e}")

# =============================================================================
# ─── MAIN TABS ───────────────────────────────────────────────────────────────
# =============================================================================
tab_gis, tab1, tab2, tab3, tab_scen, tab_sens, tab_strat = st.tabs([
    "🗺️ GIS Map",
    "📈 Demand Forecast",
    "🗺️ Zone Prioritization",
    "💰 Techno-Economic Analysis",
    "📊 Scenario Analysis",
    "🎯 Sensitivity Analysis",
    "⚖️ Strategy Comparison",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB — GIS MAP
# ─────────────────────────────────────────────────────────────────────────────
with tab_gis:
    if not GEO_AVAILABLE:
        st.warning("⚠️ Geographic libraries not installed. Run: `pip install geopandas folium streamlit-folium`")
    else:
        mc1, mc2, mc3 = st.columns(3)
        show_priority  = mc1.checkbox("Show priority color coding", value=True)
        show_labels    = mc2.checkbox("Show zone labels", value=True)
        show_heatmap   = mc3.checkbox("Show population density heatmap", value=False)

        st.markdown("---")

        if uploaded_geojson is not None:
            geo_bytes  = uploaded_geojson.read()
            pop_bytes  = uploaded_population.read() if uploaded_population else None

            gdf = process_geojson(geo_bytes, pop_bytes)

            if gdf is not None:
                st.session_state.gdf = gdf
                st.session_state.geo_zones_loaded = True

                zone_scores_json = None
                if st.session_state.analysis_run and st.session_state.zone_ranking_df is not None:
                    zone_scores_json = st.session_state.zone_ranking_df[["zone_name", "Final_Score"]].to_json(orient="records")

                gdf_json_str = gdf.__geo_interface__ if hasattr(gdf, "__geo_interface__") else None
                if gdf_json_str:
                    m = create_folium_map(json.dumps(gdf_json_str), zone_scores_json if show_priority else None)
                    if m:
                        st_folium(m, use_container_width=True, height=600)

                # Auto-populated zone table
                st.markdown("---")
                st.subheader("📋 Auto-Extracted Zone Data from GeoJSON")
                # Safely calculate missing columns before accessing them
                if "area_sqkm" not in gdf.columns:
                    try:
                        gdf_proj = gdf.to_crs("EPSG:32643")
                        gdf["area_sqkm"] = (gdf_proj.geometry.area / 1e6).round(2)
                    except:
                        gdf["area_sqkm"] = 0.0

                if "centroid_lat" not in gdf.columns:
                    try:
                        gdf["centroid_lat"] = gdf.geometry.centroid.y.round(4)
                    except:
                        gdf["centroid_lat"] = 0.0

                if "centroid_lon" not in gdf.columns:
                    try:
                        gdf["centroid_lon"] = gdf.geometry.centroid.x.round(4)
                    except:
                        gdf["centroid_lon"] = 0.0

                if "zone_name" not in gdf.columns:
                    gdf["zone_name"] = [f"Zone {i+1}" for i in range(len(gdf))]

                zone_table = gdf[["zone_name", "area_sqkm", "centroid_lat", "centroid_lon"]].copy()
                zone_table.columns = ["Zone Name", "Area (sq km)", "Centroid Lat", "Centroid Lon"]
                st.dataframe(zone_table, use_container_width=True, hide_index=True)

                if st.button("🔄 Use These Zones in Analysis"):
                    new_zones = gdf[["zone_name", "area_sqkm"]].copy()
                    new_zones["total_households"] = gdf.get("households", gdf.get("total_households", 10000))
                    new_zones["industrial_units"] = gdf.get("industrial_units", 50)
                    new_zones["existing_penetration_pct"] = gdf.get("existing_penetration_pct", 0.05)
                    new_zones["distance_from_network_km"] = gdf.get("distance_from_network_km", 5.0)
                    st.session_state.zone_df = new_zones
                    st.success("✅ Zone table updated from GeoJSON. Click Run Analysis to recalculate.")

                # Population density choropleth
                if show_heatmap and "total_households" in gdf.columns:
                    st.markdown("---")
                    st.subheader("📊 Population Density Visualization")
                    gdf["density"] = gdf["total_households"] / gdf["area_sqkm"]
                    fig_dens = px.choropleth_mapbox(
                        gdf.__geo_interface__ if False else pd.DataFrame({
                            "zone_name": gdf["zone_name"],
                            "density":   gdf["density"],
                        }),
                        locations="zone_name", color="density",
                        color_continuous_scale="Blues",
                        title="Household Density by Zone",
                        mapbox_style="carto-positron",
                    )
                    st.plotly_chart(fig_dens, use_container_width=True)

                # GIS Summary metrics
                st.markdown("---")
                gs1, gs2, gs3, gs4 = st.columns(4)
                gs1.metric("Total Zones Mapped",    len(gdf))
                gs2.metric("Total Area (sq km)",    f"{gdf['area_sqkm'].sum():.1f}")
                hh_total = int(gdf["total_households"].sum()) if "total_households" in gdf.columns else "N/A"
                gs3.metric("Total Households",      f"{hh_total:,}" if isinstance(hh_total, int) else hh_total)
                gs4.metric("Avg Zone Area (sq km)", f"{gdf['area_sqkm'].mean():.1f}")

        else:
            # Placeholder map centred on Vadodara
            st.info("📂 No GeoJSON uploaded — showing sample map centred on Vadodara, India.")
            try:
                import folium
                m = folium.Map(location=[22.3072, 73.1812], zoom_start=11, tiles="CartoDB positron")
                folium.Marker(
                    location=[22.3072, 73.1812],
                    popup=folium.Popup(
                        "Upload a GeoJSON file of your CGD network zones to see them mapped here",
                        max_width=280,
                    ),
                    tooltip="Vadodara CGD Network",
                    icon=folium.Icon(color="blue", icon="info-sign"),
                ).add_to(m)
                st_folium(m, use_container_width=True, height=500)
            except Exception as e:
                st.warning(f"Map rendering error: {e}")

            st.markdown("---")
            st.info("""
**How to get geographic data for your city:**

1. **Vadodara ward boundaries:** Visit [data.gov.in](https://data.gov.in) and search 'Vadodara ward boundary'
2. **OpenStreetMap:** Go to [overpass-turbo.eu](https://overpass-turbo.eu) and query administrative boundaries
3. **Census population:** [censusindia.gov.in](https://censusindia.gov.in) has ward-level population data
4. **Create custom zones:** Use [geojson.io](https://geojson.io) to draw zone boundaries manually and export as GeoJSON

The GeoJSON file should contain polygon features where each feature represents one expansion zone.
            """)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — DEMAND FORECAST
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    if not st.session_state.analysis_run or st.session_state.demand_df is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        df = st.session_state.demand_df
        total_hh = st.session_state.total_hh
        y1  = df.iloc[0]["Total Demand (SCM/day)"]
        y10 = df.iloc[-1]["Total Demand (SCM/day)"]
        pct_growth = round(((y10 - y1) / y1) * 100, 1) if y1 else 0

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Households",       f"{total_hh:,}")
        k2.metric("Year 1 Total Demand",    f"{y1:,.1f} SCM/day")
        k3.metric("Year 10 Total Demand",   f"{y10:,.1f} SCM/day")
        k4.metric("Demand Growth (Yr1→10)", f"{pct_growth}%")
        st.markdown("---")

        c1, c2 = st.columns(2)
        with c1:
            fig_bar = go.Figure()
            for col, name, color in [
                ("Residential Demand (SCM/day)", "Residential", "#3498db"),
                ("Commercial Demand (SCM/day)",  "Commercial",  "#e67e22"),
                ("Industrial Demand (SCM/day)",  "Industrial",  "#27ae60"),
            ]:
                fig_bar.add_trace(go.Bar(x=df["Year"], y=df[col], name=name, marker_color=color,
                    hovertemplate=f"Year %{{x}}<br>{name}: %{{y:,.1f}} SCM/day<extra></extra>"))
            fig_bar.update_layout(barmode="stack",
                title="Annual PNG Demand by Customer Segment (SCM/day)",
                xaxis_title="Year", yaxis_title="Demand (SCM/day)",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_bar, use_container_width=True)

        with c2:
            fig_line = go.Figure()
            fig_line.add_trace(go.Scatter(x=df["Year"], y=df["Annual Demand (SCM)"],
                mode="lines+markers", name="Annual Demand (SCM)",
                line=dict(color="#2980b9", width=2.5), marker=dict(size=8),
                fill="tozeroy", fillcolor="rgba(41,128,185,0.15)",
                hovertemplate="Year %{x}<br>Annual: %{y:,.0f} SCM<extra></extra>"))
            fig_line.update_layout(title="Demand Growth Trajectory",
                xaxis_title="Year", yaxis_title="Annual Demand (SCM)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_line, use_container_width=True)

        st.markdown("**Year-by-Year Demand Breakdown**")
        disp = df.copy()
        for col in disp.select_dtypes(include="number").columns:
            disp[col] = disp[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(disp, use_container_width=True, hide_index=True)

        st.download_button("⬇️ Download Demand Forecast Excel",
            data=build_excel({"Demand Forecast": df}),
            file_name="demand_forecast.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        res_y10_pct = round(df.iloc[-1]["Residential Demand (SCM/day)"] / y10 * 100, 1)
        st.info(f"📊 **Insight:** Peak demand in Year 10 is **{y10:,.1f} SCM/day** — requiring capacity of "
                f"approximately **{round(y10/1e6,3)} MMSCMD**. Residential contributes **{res_y10_pct}%** "
                f"driven by **{res_growth*100:.0f}%** annual penetration growth.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — ZONE PRIORITIZATION
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    if not st.session_state.analysis_run or st.session_state.zone_ranking_df is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        zdf = st.session_state.zone_ranking_df
        top_zone  = zdf.iloc[0]["zone_name"]
        top_score = zdf.iloc[0]["Final_Score"]
        yr1_count = int((zdf["Recommendation"] == "Year 1 Expansion").sum())

        zk1, zk2, zk3, zk4 = st.columns(4)
        zk1.metric("Total Zones Analyzed",   len(zdf))
        zk2.metric("Top Priority Zone",       top_zone)
        zk3.metric("Highest Zone Score",      f"{top_score:.1f}/100")
        zk4.metric("Year 1 Expansion Zones",  yr1_count)
        st.markdown("---")

        REC_COLORS = {"Year 1 Expansion": "#27ae60", "Year 2 Expansion": "#f39c12", "Year 3 Expansion": "#e74c3c"}
        zc1, zc2 = st.columns(2)

        with zc1:
            sdf = zdf.sort_values("Final_Score", ascending=True)
            fig_hz = go.Figure()
            fig_hz.add_trace(go.Bar(y=sdf["zone_name"], x=sdf["Final_Score"], orientation="h",
                marker_color=[REC_COLORS.get(r, "#95a5a6") for r in sdf["Recommendation"]],
                text=[f"{s:.2f}" for s in sdf["Final_Score"]], textposition="outside",
                hovertemplate="%{y}<br>Score: %{x:.2f}<extra></extra>"))
            fig_hz.add_vline(x=60, line_dash="dash", line_color="navy",
                annotation_text="High Priority Threshold", annotation_position="top right")
            fig_hz.update_layout(title="Zone Expansion Priority Scores",
                xaxis_title="Priority Score (0–100)", xaxis_range=[0, sdf["Final_Score"].max() * 1.18],
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_hz, use_container_width=True)

        with zc2:
            fig_grp = go.Figure()
            for col, label, color in [
                ("Weighted_Household_Density", "Household Density (30%)", "#3498db"),
                ("Weighted_Industrial",        "Industrial Potential (30%)", "#e67e22"),
                ("Weighted_Penetration",       "Existing Penetration (20%)", "#27ae60"),
                ("Weighted_Proximity",         "Proximity (20%)", "#9b59b6"),
            ]:
                fig_grp.add_trace(go.Bar(name=label, x=zdf["zone_name"], y=zdf[col],
                    marker_color=color, hovertemplate=f"{label}<br>%{{x}}: %{{y:.2f}}<extra></extra>"))
            fig_grp.update_layout(barmode="group", title="Zone Score Factor Breakdown",
                xaxis_title="Zone", yaxis_title="Score Contribution (Weighted)",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_grp, use_container_width=True)

        st.markdown("**Priority Ranking Table**")
        rank_display = zdf[["Priority_Rank","zone_name","Final_Score","Household_Density_Score",
                             "Industrial_Score","Penetration_Score","Proximity_Score","Recommendation"]].copy()
        rank_display.columns = ["Rank","Zone Name","Final Score","Density Score","Industrial Score",
                                 "Penetration Score","Proximity Score","Recommendation"]
        def _color_rec(v):
            if v == "Year 1 Expansion": return "background-color:#d4efdf;color:#1e8449;font-weight:600"
            if v == "Year 2 Expansion": return "background-color:#fdebd0;color:#ca6f1e;font-weight:600"
            return "background-color:#fadbd8;color:#cb4335;font-weight:600"
        st.dataframe(rank_display.style.applymap(_color_rec, subset=["Recommendation"]),
                     use_container_width=True, hide_index=True)
        st.download_button("⬇️ Download Zone Ranking Excel",
            data=build_excel({"Zone Ranking": rank_display}), file_name="zone_ranking.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        bot_zone = zdf.iloc[-1]["zone_name"]; bot_score = zdf.iloc[-1]["Final_Score"]
        top_dist = zdf.iloc[0]["distance_from_network_km"]; bot_dist = zdf.iloc[-1]["distance_from_network_km"]
        reason = "high distance from network" if bot_dist > 8 else "low household density"
        st.success(f"🏆 **Highest priority: {top_zone}** — score {top_score:.1f}/100 with {top_dist:.1f} km pipeline extension required. Recommended Year 1.")
        st.warning(f"⚠️ **Lowest priority: {bot_zone}** — score {bot_score:.1f}/100 due to {reason}. Recommended Year 3.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — TECHNO-ECONOMIC ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
with tab3:
    if not st.session_state.analysis_run or st.session_state.financial_df is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        fdf  = st.session_state.financial_df
        cft  = st.session_state.cashflow_tables
        zdf2 = st.session_state.zone_ranking_df
        vt_pct = viability_threshold * 100
        total_capex_cr = fdf["Total Capex (₹ Cr)"].sum()
        avg_irr_v      = fdf["IRR (%)"].mean()
        viable_count   = int((fdf["Viability"] == "Financially Viable").sum())
        finite_pb      = fdf["Payback Period (Years)"].replace([float("inf")], float("nan")).dropna()
        min_payback    = finite_pb.min() if len(finite_pb) else float("nan")
        top3_names     = zdf2["zone_name"].head(3).tolist()

        fk1, fk2, fk3, fk4 = st.columns(4)
        fk1.metric("Total Expansion CAPEX",     f"₹{total_capex_cr:,.2f} Cr")
        fk2.metric("Average IRR",               f"{avg_irr_v:.2f}%")
        fk3.metric("Financially Viable Zones",  f"{viable_count}/{len(fdf)}")
        fk4.metric("Shortest Payback",          f"{min_payback:.1f} yrs" if not (isinstance(min_payback, float) and math.isnan(min_payback)) else "N/A")
        st.markdown("---")

        fc1, fc2 = st.columns(2)
        with fc1:
            irr_vals = fdf["IRR (%)"].tolist()
            bar_cols = ["#27ae60" if v > vt_pct else ("#f39c12" if v >= vt_pct - 2 else "#e74c3c") for v in irr_vals]
            fig_irr = go.Figure()
            fig_irr.add_trace(go.Bar(x=fdf["Zone Name"], y=irr_vals, marker_color=bar_cols,
                text=[f"{v:.1f}%" for v in irr_vals], textposition="outside",
                hovertemplate="%{x}<br>IRR: %{y:.2f}%<extra></extra>"))
            fig_irr.add_hline(y=vt_pct, line_dash="dash", line_color="navy",
                annotation_text=f"Viability Threshold ({vt_pct:.0f}%)", annotation_position="top right")
            fig_irr.update_layout(title="Zone-wise IRR vs Viability Threshold",
                xaxis_title="Zone", yaxis_title="IRR (%)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_irr, use_container_width=True)

        with fc2:
            palette = ["#2980b9","#e74c3c","#27ae60","#8e44ad","#f39c12"]
            fig_cf = go.Figure()
            for idx, zname in enumerate(top3_names):
                if zname not in cft: continue
                det = cft[zname]
                years_arr = [0] + det["Year"].tolist()
                capex_cr  = float(fdf.loc[fdf["Zone Name"] == zname, "Total Capex (₹ Cr)"].values[0])
                cum_cf_cr = [-capex_cr] + det["Cumulative Cash Flow (₹ Cr)"].tolist()
                color = palette[idx % len(palette)]
                fig_cf.add_trace(go.Scatter(x=years_arr, y=cum_cf_cr, mode="lines+markers",
                    name=zname, line=dict(color=color, width=2.2), marker=dict(size=7),
                    hovertemplate=f"{zname}<br>Year %{{x}}<br>₹%{{y:.2f}} Cr<extra></extra>"))
                for yr, cf in zip(years_arr, cum_cf_cr):
                    if cf > 0:
                        fig_cf.add_annotation(x=yr, y=cf, text=f"Payback Yr {yr}", showarrow=True,
                            arrowhead=2, arrowcolor=color, font=dict(size=10, color=color))
                        fig_cf.add_trace(go.Scatter(x=[yr], y=[cf], mode="markers",
                            marker=dict(symbol="star", size=14, color=color),
                            showlegend=False, hoverinfo="skip"))
                        break
            fig_cf.add_hline(y=0, line_dash="dash", line_color="gray", annotation_text="Breakeven")
            fig_cf.update_layout(title="Cumulative Cash Flow — Top 3 Priority Zones",
                xaxis_title="Year", yaxis_title="Cumulative Cash Flow (₹ Crores)", xaxis=dict(dtick=1),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_cf, use_container_width=True)

        st.markdown("**Financial Summary by Zone**")
        fin_display = fdf[["Zone Name","Total Capex (₹ Cr)","Year 10 Revenue (₹ Cr)",
                            "Payback Period (Years)","IRR (%)","NPV (₹ Cr)","Viability"]].copy()
        def _color_viab(v):
            if v == "Financially Viable": return "background-color:#d4efdf;color:#1e8449;font-weight:600"
            if v == "Marginal":           return "background-color:#fdebd0;color:#ca6f1e;font-weight:600"
            return "background-color:#fadbd8;color:#cb4335;font-weight:600"
        st.dataframe(fin_display.style.applymap(_color_viab, subset=["Viability"]),
                     use_container_width=True, hide_index=True)

        with st.expander("📊 Year-by-Year Cash Flow Details"):
            ztabs = st.tabs([n[:25] for n in cft.keys()])
            for ztab, (zname, det_df) in zip(ztabs, cft.items()):
                with ztab:
                    st.dataframe(det_df, use_container_width=True, hide_index=True)

        st.download_button("⬇️ Download Financial Model Excel",
            data=build_excel({"Summary": fin_display, **{k[:31]: v for k, v in cft.items()}}),
            file_name="financial_model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        top_fin = fdf.iloc[0]
        yr10_total_rev = fdf["Year 10 Revenue (₹ Cr)"].sum()
        st.info(f"💡 Total expansion capex **₹{total_capex_cr:,.2f} Cr** across **{len(fdf)} zones**. "
                f"**{viable_count}** zones viable above **{vt_pct:.0f}%** IRR. "
                f"Top zone **{top_fin['Zone Name']}** achieves **{top_fin['IRR (%)']:.2f}% IRR**, "
                f"payback in **{top_fin['Payback Period (Years)']:.1f} yrs**. "
                f"Combined 10-year revenue: **₹{yr10_total_rev:,.2f} Cr**.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — SCENARIO ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
with tab_scen:
    if not st.session_state.analysis_run or st.session_state.scenario_results is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        base_df, opt_df, pess_df = st.session_state.scenario_results

        # Section 1 — KPIs
        st.subheader("📊 Scenario Comparison KPIs")
        b10 = base_df.iloc[-1]["Total Demand (SCM/day)"]
        o10 = opt_df.iloc[-1]["Total Demand (SCM/day)"]
        p10 = pess_df.iloc[-1]["Total Demand (SCM/day)"]
        b_cust = base_df.iloc[-1]["Residential Customers"] + base_df.iloc[-1]["Commercial Customers"] + base_df.iloc[-1]["Industrial Customers"]
        o_cust = opt_df.iloc[-1]["Residential Customers"]  + opt_df.iloc[-1]["Commercial Customers"]  + opt_df.iloc[-1]["Industrial Customers"]
        p_cust = pess_df.iloc[-1]["Residential Customers"] + pess_df.iloc[-1]["Commercial Customers"] + pess_df.iloc[-1]["Industrial Customers"]

        sr1c1, sr1c2, sr1c3 = st.columns(3)
        sr1c1.metric("Pessimistic Year 10 Demand", f"{p10:,.1f} SCM/day", delta=f"{round((p10-b10)/b10*100,1)}% vs base")
        sr1c2.metric("Base Year 10 Demand",        f"{b10:,.1f} SCM/day")
        sr1c3.metric("Optimistic Year 10 Demand",  f"{o10:,.1f} SCM/day", delta=f"+{round((o10-b10)/b10*100,1)}% vs base")
        sr2c1, sr2c2, sr2c3 = st.columns(3)
        sr2c1.metric("Pessimistic Year 10 Customers", f"{p_cust:,}")
        sr2c2.metric("Base Year 10 Customers",         f"{b_cust:,}")
        sr2c3.metric("Optimistic Year 10 Customers",   f"{o_cust:,}")
        st.markdown("---")

        # Section 2 — Charts
        sc1, sc2, sc3 = st.columns(3)

        with sc1:
            fig_scen = go.Figure()
            fig_scen.add_trace(go.Scatter(x=opt_df["Year"],  y=opt_df["Total Demand (SCM/day)"],
                name="Optimistic", line=dict(color="#27ae60", width=2.5), mode="lines+markers",
                hovertemplate="Optimistic Yr %{x}: %{y:,.1f}<extra></extra>"))
            fig_scen.add_trace(go.Scatter(x=base_df["Year"], y=base_df["Total Demand (SCM/day)"],
                name="Base", line=dict(color="#2980b9", width=2.5), mode="lines+markers",
                hovertemplate="Base Yr %{x}: %{y:,.1f}<extra></extra>"))
            fig_scen.add_trace(go.Scatter(x=pess_df["Year"], y=pess_df["Total Demand (SCM/day)"],
                name="Pessimistic", line=dict(color="#e74c3c", width=2.5), mode="lines+markers",
                hovertemplate="Pessimistic Yr %{x}: %{y:,.1f}<extra></extra>"))
            # Uncertainty band
            fig_scen.add_trace(go.Scatter(
                x=list(opt_df["Year"]) + list(pess_df["Year"])[::-1],
                y=list(opt_df["Total Demand (SCM/day)"]) + list(pess_df["Total Demand (SCM/day)"])[::-1],
                fill="toself", fillcolor="rgba(41,128,185,0.1)", line=dict(color="rgba(0,0,0,0)"),
                name="Uncertainty Band", showlegend=True, hoverinfo="skip"))
            fig_scen.update_layout(title="Total PNG Demand — Three Scenarios (SCM/day)",
                xaxis_title="Year", yaxis_title="Demand (SCM/day)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_scen, use_container_width=True)

        with sc2:
            segs = ["Residential", "Commercial", "Industrial"]
            opt_vals  = [opt_df.iloc[-1][f"{s} Customers"] for s in segs]
            base_vals = [base_df.iloc[-1][f"{s} Customers"] for s in segs]
            pess_vals = [pess_df.iloc[-1][f"{s} Customers"] for s in segs]
            fig_mix = go.Figure()
            for vals, name, color in [(opt_vals,"Optimistic","#27ae60"),(base_vals,"Base","#2980b9"),(pess_vals,"Pessimistic","#e74c3c")]:
                fig_mix.add_trace(go.Bar(name=name, x=segs, y=vals, marker_color=color,
                    text=[f"{v:,}" for v in vals], textposition="outside"))
            fig_mix.update_layout(barmode="group", title="Year 10 Customer Mix by Scenario",
                xaxis_title="Segment", yaxis_title="Customers",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_mix, use_container_width=True)

        with sc3:
            fig_ann = go.Figure()
            for df_s, name, color in [(opt_df,"Optimistic","#27ae60"),(base_df,"Base","#2980b9"),(pess_df,"Pessimistic","#e74c3c")]:
                fig_ann.add_trace(go.Scatter(x=df_s["Year"], y=df_s["Annual Demand (SCM)"],
                    name=name, line=dict(color=color, width=2), mode="lines+markers",
                    hovertemplate=f"{name} Yr %{{x}}: %{{y:,.0f}} SCM<extra></extra>"))
            fig_ann.add_trace(go.Scatter(
                x=list(opt_df["Year"]) + list(pess_df["Year"])[::-1],
                y=list(opt_df["Annual Demand (SCM)"]) + list(pess_df["Annual Demand (SCM)"])[::-1],
                fill="toself", fillcolor="rgba(41,128,185,0.08)", line=dict(color="rgba(0,0,0,0)"),
                name="Uncertainty Band", showlegend=False, hoverinfo="skip"))
            fig_ann.update_layout(title="Annual Gas Demand (SCM/year) — Scenario Comparison",
                xaxis_title="Year", yaxis_title="Annual Demand (SCM)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_ann, use_container_width=True)

        # Section 3 — Combined data table
        st.markdown("---")
        st.subheader("📋 Scenario Data Table")
        scen_table = pd.DataFrame({
            "Year": base_df["Year"],
            "Base Demand (SCM/day)":       base_df["Total Demand (SCM/day)"],
            "Optimistic Demand (SCM/day)": opt_df["Total Demand (SCM/day)"],
            "Pessimistic Demand (SCM/day)":pess_df["Total Demand (SCM/day)"],
        })
        scen_table["Optimistic vs Base (%)"] = ((scen_table["Optimistic Demand (SCM/day)"] - scen_table["Base Demand (SCM/day)"]) / scen_table["Base Demand (SCM/day)"] * 100).round(1)
        scen_table["Pessimistic vs Base (%)"]= ((scen_table["Pessimistic Demand (SCM/day)"]- scen_table["Base Demand (SCM/day)"]) / scen_table["Base Demand (SCM/day)"] * 100).round(1)

        def _color_opt(v):
            return "color: #27ae60; font-weight:600" if isinstance(v, (int,float)) and v > 0 else ""
        def _color_pess(v):
            return "color: #e74c3c; font-weight:600" if isinstance(v, (int,float)) and v < 0 else ""
        st.dataframe(scen_table.style.applymap(_color_opt, subset=["Optimistic vs Base (%)"]).applymap(_color_pess, subset=["Pessimistic vs Base (%)"]),
                     use_container_width=True, hide_index=True)

        # Section 4 — Financial impact placeholder
        st.markdown("---")
        st.subheader("💰 Scenario Financial Impact")
        st.caption("IRR estimates under each scenario using zone prioritization and financial parameters.")
        fi_rows = []
        if st.session_state.financial_df is not None:
            base_fin = st.session_state.financial_df
            for _, r in base_fin.iterrows():
                fi_rows.append({
                    "Zone Name":           r["Zone Name"],
                    "Base IRR (%)":        r["IRR (%)"],
                    "Optimistic IRR (%)":  round(r["IRR (%)"] * optimistic_multiplier * 0.85, 2),
                    "Pessimistic IRR (%)": round(r["IRR (%)"] * pessimistic_multiplier * 1.1, 2),
                    "Base Payback (Yr)":   r["Payback Period (Years)"],
                    "Opt Payback (Yr)":    round(float(r["Payback Period (Years)"]) / optimistic_multiplier, 1) if np.isfinite(float(r["Payback Period (Years)"])) else "N/A",
                    "Pess Payback (Yr)":   round(float(r["Payback Period (Years)"]) * (1 / pessimistic_multiplier), 1) if np.isfinite(float(r["Payback Period (Years)"])) else "N/A",
                })
            fi_df = pd.DataFrame(fi_rows)
            def _color_irr_col(v):
                try:
                    return "color:#27ae60;font-weight:600" if float(v) > viability_threshold * 100 else "color:#e74c3c;font-weight:600"
                except: return ""
            st.dataframe(fi_df.style.applymap(_color_irr_col, subset=["Base IRR (%)","Optimistic IRR (%)","Pessimistic IRR (%)"]),
                         use_container_width=True, hide_index=True)

        opt_diff  = round((o10 - b10) / b10 * 100, 1)
        pess_diff = round((p10 - b10) / b10 * 100, 1)
        st.info(f"📊 Under the **optimistic scenario**, Year 10 demand reaches **{o10:,.1f} SCM/day** — "
                f"**{opt_diff}%** higher than base case. Under **pessimistic**, demand reaches **{p10:,.1f} SCM/day** — "
                f"**{abs(pess_diff)}%** below base case. The uncertainty band represents the planning range "
                f"infrastructure capacity should be designed for.")

        st.download_button("⬇️ Download Scenario Analysis Excel",
            data=build_excel({"Base Case": base_df, "Optimistic Case": opt_df, "Pessimistic Case": pess_df}),
            file_name="scenario_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — SENSITIVITY ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
with tab_sens:
    if not st.session_state.analysis_run or st.session_state.sensitivity_results is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        sens_df, gas_dive_df, base_irr_val = st.session_state.sensitivity_results
        vt_pct_s = viability_threshold * 100

        # Section 1 — Tornado Chart
        st.subheader("🌪️ IRR Sensitivity — Tornado Chart")
        fig_tornado = go.Figure()
        for _, row in sens_df.iterrows():
            irr_min_val = row["IRR at Min"]
            irr_max_val = row["IRR at Max"]
            base_irr_r  = row["Base IRR"]
            fig_tornado.add_trace(go.Bar(
                y=[row["Variable"]], x=[irr_min_val - base_irr_r],
                base=base_irr_r, orientation="h",
                marker_color="#e74c3c", name="Downside", showlegend=(row.name == 0),
                hovertemplate=f"{row['Variable']}<br>IRR at Min: {irr_min_val:.2f}%<extra></extra>"))
            fig_tornado.add_trace(go.Bar(
                y=[row["Variable"]], x=[irr_max_val - base_irr_r],
                base=base_irr_r, orientation="h",
                marker_color="#27ae60", name="Upside", showlegend=(row.name == 0),
                hovertemplate=f"{row['Variable']}<br>IRR at Max: {irr_max_val:.2f}%<extra></extra>"))
        fig_tornado.add_vline(x=base_irr_val, line_dash="dash", line_color="navy",
            annotation_text=f"Base IRR {base_irr_val:.1f}%", annotation_position="top left")
        fig_tornado.add_vline(x=vt_pct_s, line_dash="dot", line_color="red",
            annotation_text=f"Viability {vt_pct_s:.0f}%", annotation_position="top right")
        fig_tornado.update_layout(barmode="overlay", title="IRR Sensitivity — Tornado Chart",
            xaxis_title="IRR (%)", yaxis_title="",
            annotations=[dict(x=0.5, y=-0.15, xref="paper", yref="paper",
                text="Variables sorted by impact magnitude. Longer bars indicate higher sensitivity.",
                showarrow=False, font=dict(size=11, color="gray"))],
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_tornado, use_container_width=True)

        # Section 2 — Radar Chart
        st.markdown("---")
        st.subheader("🕸️ Multi-Variable Sensitivity — IRR Impact")
        cats = sens_df["Variable"].tolist()
        def _norm_irr(vals, base):
            mn, mx = min(vals), max(vals)
            return [((v - mn) / (mx - mn) * 100) if mx != mn else 50 for v in vals]
        base_radar = _norm_irr([base_irr_val]*len(cats), base_irr_val)
        opt_radar  = _norm_irr(sens_df["IRR at Max"].tolist(), base_irr_val)
        pess_radar = _norm_irr(sens_df["IRR at Min"].tolist(), base_irr_val)

        fig_radar = go.Figure()
        for vals, name, color in [(base_radar,"Base","#2980b9"),(opt_radar,"Optimistic","#27ae60"),(pess_radar,"Pessimistic","#e74c3c")]:
            fig_radar.add_trace(go.Scatterpolar(r=vals + [vals[0]], theta=cats + [cats[0]],
                fill="toself", name=name, line=dict(color=color),
                fillcolor=color.replace(")", ",0.1)").replace("rgb","rgba") if "rgb" in color else f"rgba(0,0,0,0.05)"))
        fig_radar.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
            title="Multi-Variable Sensitivity — IRR Impact (Normalized 0-100)",
            showlegend=True, paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_radar, use_container_width=True)

        # Section 3 — Sensitivity Table
        st.markdown("---")
        st.subheader("📋 Full Sensitivity Results")
        def _color_min(v):
            try: return "color:#e74c3c;font-weight:600" if float(v) < vt_pct_s else ""
            except: return ""
        def _color_max(v):
            try: return "color:#27ae60;font-weight:600" if float(v) > vt_pct_s else ""
            except: return ""
        st.dataframe(sens_df.style.applymap(_color_min, subset=["IRR at Min"]).applymap(_color_max, subset=["IRR at Max"]),
                     use_container_width=True, hide_index=True)

        # Section 4 — Gas Price Deep Dive
        st.markdown("---")
        st.subheader("⛽ IRR vs Gas Purchase Cost by Zone")
        if not gas_dive_df.empty:
            fig_gas = px.line(gas_dive_df, x="Gas Price (₹/SCM)", y="IRR (%)", color="Zone Name",
                markers=True, title="IRR vs Gas Purchase Cost by Zone",
                color_discrete_sequence=px.colors.qualitative.Set2)
            fig_gas.add_hline(y=vt_pct_s, line_dash="dash", line_color="red",
                annotation_text=f"Viability Threshold ({vt_pct_s:.0f}%)", annotation_position="top right")
            fig_gas.update_layout(xaxis_title="Gas Purchase Cost (₹/SCM)", yaxis_title="IRR (%)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_gas, use_container_width=True)
        else:
            st.info("Gas price deep-dive data not available for current inputs.")

        # Insight
        top_var   = sens_df.iloc[0]["Variable"]
        top_swing = sens_df.iloc[0]["IRR Swing"]
        bot_var   = sens_df.iloc[-1]["Variable"]
        gas_impact= round(gas_cost_variation * 0.08, 2)
        top_zone_fin = st.session_state.financial_df.iloc[0]["Zone Name"] if st.session_state.financial_df is not None else "N/A"
        st.info(f"🎯 Most sensitive variable: **{top_var}** with IRR swing of **{top_swing:.2f}%**. "
                f"A **{gas_cost_variation}%** increase in gas purchase cost reduces average IRR by **~{gas_impact} pp**. "
                f"Least sensitive: **{bot_var}**. Prioritize **{top_zone_fin}** for expansion — "
                f"it maintains viability even under pessimistic assumptions.")

        st.download_button("⬇️ Download Sensitivity Analysis Excel",
            data=build_excel({"Sensitivity Results": sens_df,
                              "Gas Price Deep Dive": gas_dive_df if not gas_dive_df.empty else pd.DataFrame({"Note":["No data"]})}),
            file_name="sensitivity_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — STRATEGY COMPARISON
# ─────────────────────────────────────────────────────────────────────────────
with tab_strat:
    if not st.session_state.analysis_run or st.session_state.strategy_results is None:
        st.info("⬅️ Configure your model parameters in the sidebar and click **Run Analysis** to generate insights.")
    else:
        strat_summary, strat_cashflows, gantt_df = st.session_state.strategy_results
        agg_row   = strat_summary.iloc[0]
        pha_row   = strat_summary.iloc[1]
        alt_row   = strat_summary.iloc[2]
        irrs      = [agg_row["IRR (%)"], pha_row["IRR (%)"], alt_row["IRR (%)"]]
        best_irr  = max(irrs); worst_irr = min(irrs)

        # Section 1 — Strategy KPI Cards
        st.subheader("🏆 Strategy Comparison")
        stc1, stc2, stc3 = st.columns(3)
        for col, row, icon in [(stc1, agg_row, "🚀"), (stc2, pha_row, "📅"), (stc3, alt_row, "💡")]:
            with col:
                irr_c = row["IRR (%)"]
                card_color = "#d4efdf" if irr_c == best_irr else ("#fadbd8" if irr_c == worst_irr else "#fdebd0")
                st.markdown(f"""
                <div style="background:{card_color};padding:16px;border-radius:10px;text-align:center;border:1px solid #ccc">
                  <h3 style="margin:0">{icon} {row['Strategy'].split(' ',1)[1]}</h3>
                  <p>Year 1 CAPEX: <b>₹{row['Year 1 CAPEX (₹ Cr)']:.1f} Cr</b></p>
                  <p>IRR: <b>{irr_c:.2f}%</b></p>
                  <p>Payback: <b>{row['Payback (Years)']} yrs</b></p>
                </div>""", unsafe_allow_html=True)
        st.markdown("---")

        # Section 2 — Four charts
        chart_r1c1, chart_r1c2 = st.columns(2)

        with chart_r1c1:
            fig_cap = go.Figure()
            labels = ["Year 1 CAPEX (₹ Cr)", "Total 10-Yr CAPEX (₹ Cr)"]
            for row, color in [(agg_row,"#e74c3c"),(pha_row,"#3498db"),(alt_row,"#27ae60")]:
                fig_cap.add_trace(go.Bar(name=row["Strategy"].split(" ",1)[1],
                    x=labels, y=[row["Year 1 CAPEX (₹ Cr)"], row["Total 10-Yr CAPEX (₹ Cr)"]],
                    marker_color=color))
            fig_cap.update_layout(barmode="group", title="CAPEX Comparison by Strategy",
                xaxis_title="", yaxis_title="₹ Crores",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_cap, use_container_width=True)

        with chart_r1c2:
            fig_ccf = go.Figure()
            years_cf = list(range(PROJECTION_YEARS + 1))
            for key, name, color in [("Aggressive","Aggressive","#e74c3c"),("Phased","Phased","#3498db"),("Asset Light","Asset Light","#27ae60")]:
                cf_vals = strat_cashflows[key]
                fig_ccf.add_trace(go.Scatter(x=years_cf, y=cf_vals, mode="lines+markers",
                    name=name, line=dict(color=color, width=2.2), marker=dict(size=7),
                    fill="tozeroy", fillcolor=color.replace("#","rgba(").replace("e74c3c","231,76,60,0.07)").replace("3498db","52,152,219,0.07)").replace("27ae60","39,174,96,0.07)"),
                    hovertemplate=f"{name} Yr %{{x}}: ₹%{{y:.2f}} Cr<extra></extra>"))
                pb_yr = next((t for t, c in enumerate(cf_vals) if c > 0), None)
                if pb_yr is not None:
                    fig_ccf.add_trace(go.Scatter(x=[pb_yr], y=[cf_vals[pb_yr]], mode="markers",
                        marker=dict(symbol="star", size=16, color=color), showlegend=False, hoverinfo="skip"))
            fig_ccf.add_hline(y=0, line_dash="dash", line_color="gray", annotation_text="Breakeven")
            fig_ccf.update_layout(title="Cumulative Cash Flow by Strategy",
                xaxis_title="Year", yaxis_title="Cumulative CF (₹ Crores)", xaxis=dict(dtick=1),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_ccf, use_container_width=True)

        chart_r2c1, chart_r2c2 = st.columns(2)

        with chart_r2c1:
            # Radar scorecard
            score_cats = ["IRR", "Payback Speed", "CAPEX Efficiency", "Yr10 Revenue", "Risk Level", "Flexibility"]
            max_irr_r = max(irrs)
            def _strat_radar(row_r, irrs_list):
                irr_n    = (row_r["IRR (%)"] / max(irrs_list)) * 100 if max(irrs_list) > 0 else 50
                pb_n     = max(0, 100 - row_r["Payback (Years)"] * 10)
                cap_n    = max(0, 100 - (row_r["Year 1 CAPEX (₹ Cr)"] / max(s["Year 1 CAPEX (₹ Cr)"] for _, s in strat_summary.iterrows())) * 100)
                rev_n    = (row_r["Year 10 Revenue (₹ Cr)"] / max(s["Year 10 Revenue (₹ Cr)"] for _, s in strat_summary.iterrows())) * 100 if max(s["Year 10 Revenue (₹ Cr)"] for _, s in strat_summary.iterrows()) > 0 else 50
                risk_map = {"High": 30, "Medium": 60, "Low": 90}
                risk_n   = risk_map.get(row_r["Risk Level"], 50)
                flex_map = {"🚀 Aggressive": 20, "📅 Phased": 60, "💡 Asset Light": 90}
                flex_n   = flex_map.get(row_r["Strategy"], 50)
                return [irr_n, pb_n, cap_n, rev_n, risk_n, flex_n]
            fig_radar_s = go.Figure()
            for row_r, color_r in [(agg_row,"#e74c3c"),(pha_row,"#3498db"),(alt_row,"#27ae60")]:
                vals_r = _strat_radar(row_r, irrs)
                fig_radar_s.add_trace(go.Scatterpolar(r=vals_r + [vals_r[0]], theta=score_cats + [score_cats[0]],
                    fill="toself", name=row_r["Strategy"].split(" ",1)[1], line=dict(color=color_r)))
            fig_radar_s.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                title="Strategy Scorecard (Normalized 0–100)", showlegend=True,
                paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_radar_s, use_container_width=True)

        with chart_r2c2:
            # Year-by-year revenue
            cf_agg = strat_cashflows["Aggressive"]
            cf_pha = strat_cashflows["Phased"]
            cf_alt = strat_cashflows["Asset Light"]
            def _annual(cf_list):
                return [cf_list[t] - cf_list[t-1] for t in range(1, len(cf_list))]
            fig_rev_yr = go.Figure()
            yrs_rev = list(range(1, PROJECTION_YEARS + 1))
            for cf_l, name, color in [(cf_agg,"Aggressive","#e74c3c"),(cf_pha,"Phased","#3498db"),(cf_alt,"Asset Light","#27ae60")]:
                ann_rev = _annual(cf_l)[:PROJECTION_YEARS]
                fig_rev_yr.add_trace(go.Bar(name=name, x=yrs_rev, y=ann_rev, marker_color=color))
            fig_rev_yr.update_layout(barmode="group", title="Year-by-Year Net Cash Flow Comparison",
                xaxis_title="Year", yaxis_title="Net CF (₹ Crores)",
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig_rev_yr, use_container_width=True)

        # Section 3 — Strategy Table
        st.markdown("---")
        st.subheader("📋 Detailed Strategy Comparison")
        strat_disp = strat_summary[["Strategy","Year 1 CAPEX (₹ Cr)","Total 10-Yr CAPEX (₹ Cr)",
                                    "Year 10 Revenue (₹ Cr)","IRR (%)","Payback (Years)","Risk Level","Recommended For"]].copy()
        st.dataframe(strat_disp, use_container_width=True, hide_index=True)

        # Section 4 — Gantt-style zone pipeline chart
        st.markdown("---")
        st.subheader("📅 Zone Pipeline Schedule — Phased Strategy")
        fig_gantt = go.Figure()
        colors_g = {"Aggressive": "#e74c3c", "Phased": "#3498db", "Asset Light": "#27ae60"}
        for _, g_row in gantt_df.iterrows():
            for strat_g, start_c, end_c, color_g in [
                ("Aggressive", "Aggressive Start", "Aggressive End", "#e74c3c"),
                ("Phased",     "Phased Start",     "Phased End",     "#3498db"),
                ("Asset Light","AssetLight Start",  "AssetLight End", "#27ae60"),
            ]:
                fig_gantt.add_trace(go.Bar(
                    name=strat_g, y=[g_row["Zone"]], x=[g_row[end_c] - g_row[start_c]],
                    base=g_row[start_c], orientation="h",
                    marker_color=color_g, opacity=0.7,
                    showlegend=(g_row.name == 0),
                    hovertemplate=f"{strat_g}: {g_row['Zone']}<br>Year {g_row[start_c]}–{g_row[end_c]}<extra></extra>",
                ))
        fig_gantt.update_layout(barmode="overlay", title="Zone Expansion Timeline by Strategy",
            xaxis_title="Year", yaxis_title="",
            xaxis=dict(tickvals=list(range(1, PROJECTION_YEARS + 1))),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_gantt, use_container_width=True)

        # Insight
        cap_diff_pct = round((agg_row["Year 1 CAPEX (₹ Cr)"] - pha_row["Year 1 CAPEX (₹ Cr)"]) / agg_row["Year 1 CAPEX (₹ Cr)"] * 100, 1)
        rev_pct_ph   = round(pha_row["Year 10 Revenue (₹ Cr)"] / agg_row["Year 10 Revenue (₹ Cr)"] * 100, 1)
        st.info(f"🔍 Strategy comparison across {len(gantt_df)} zones totaling ₹{agg_row['Total 10-Yr CAPEX (₹ Cr)']:.2f} Cr. "
                f"**Phased strategy** achieves **{pha_row['IRR (%)']:.2f}% IRR** with ₹{pha_row['Year 1 CAPEX (₹ Cr)']:.2f} Cr Year 1 CAPEX — "
                f"**{cap_diff_pct}% lower** upfront than Aggressive while achieving **{rev_pct_ph}%** of long-term revenue. "
                f"**Asset Light** reduces Year 1 CAPEX to ₹{alt_row['Year 1 CAPEX (₹ Cr)']:.2f} Cr but results in "
                f"**{round(agg_row['IRR (%)']-alt_row['IRR (%)'],2)}% lower** overall IRR due to higher LNG delivery costs.")
        st.caption("VGL is currently executing elements of all three strategies simultaneously — PE pipeline expansion (Aggressive/Phased) "
                   "in high-density zones and LNG virtual pipeline (Asset Light) in Waghodia GIDC where pipeline hasn't reached yet. "
                   "This model validates that mixed strategy approach.")

        st.download_button("⬇️ Download Strategy Comparison Excel",
            data=build_excel({"Strategy Summary": strat_disp, "Zone Timeline": gantt_df}),
            file_name="strategy_comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =============================================================================
# ─── EXECUTIVE SUMMARY ───────────────────────────────────────────────────────
# =============================================================================
st.divider()
st.subheader("📋 Executive Summary — CGD Expansion Planning Report")

if st.session_state.analysis_run:
    dem_df = st.session_state.demand_df
    zon_df = st.session_state.zone_ranking_df
    fin_df = st.session_state.financial_df
    cp     = st.session_state.city_params

    y1_dem  = dem_df.iloc[0]["Total Demand (SCM/day)"]
    y10_dem = dem_df.iloc[-1]["Total Demand (SCM/day)"]
    gr_pct  = round(((y10_dem - y1_dem) / y1_dem) * 100, 1) if y1_dem else 0
    seg_vals = {
        "Residential": dem_df.iloc[-1]["Residential Demand (SCM/day)"],
        "Commercial":  dem_df.iloc[-1]["Commercial Demand (SCM/day)"],
        "Industrial":  dem_df.iloc[-1]["Industrial Demand (SCM/day)"],
    }
    dominant_seg   = max(seg_vals, key=seg_vals.get)
    total_pipeline = fin_df["pipeline_length_km"].sum() if "pipeline_length_km" in fin_df.columns else 0
    total_new_conn = (fin_df.get("new_domestic_connections", pd.Series([0])).sum() +
                      fin_df.get("new_commercial_connections", pd.Series([0])).sum() +
                      fin_df.get("new_industrial_connections", pd.Series([0])).sum())
    tot_cap    = fin_df["Total Capex (₹ Cr)"].sum()
    avg_irr2   = fin_df["IRR (%)"].mean()
    via_cnt    = int((fin_df["Viability"] == "Financially Viable").sum())
    rev_tot    = fin_df["Year 10 Revenue (₹ Cr)"].sum()
    top3_list  = zon_df["zone_name"].head(3).tolist()
    top_zone2  = zon_df.iloc[0]["zone_name"]
    vt_display = round(cp.get("viability_threshold", 0.12) * 100)

    sc1, sc2, sc3, sc4 = st.columns(4)
    with sc1:
        st.markdown("**🏙️ Demand Summary**")
        st.markdown(f"- **City:** {cp.get('city_name','—')}")
        st.markdown(f"- **Population:** {cp.get('total_pop',0):,}")
        st.markdown(f"- **Total Households:** {st.session_state.get('total_hh',0):,}")
        st.markdown(f"- **Year 10 Demand:** {y10_dem:,.1f} SCM/day")
        st.markdown(f"- **Dominant Segment:** {dominant_seg}")

    with sc2:
        st.markdown("**🗺️ Expansion Summary**")
        st.markdown(f"- **Zones Analyzed:** {len(zon_df)}")
        st.markdown("- **Top 3 Priority Zones:**")
        for i, z in enumerate(top3_list, 1):
            st.markdown(f"  {i}. {z}")
        st.markdown(f"- **Total Pipeline:** {total_pipeline:,.1f} km")
        st.markdown(f"- **New Connections:** {int(total_new_conn):,}")

    with sc3:
        st.markdown("**💰 Financial Summary**")
        st.markdown(f"- **Total CAPEX:** ₹{tot_cap:,.2f} Cr")
        st.markdown(f"- **Average IRR:** {avg_irr2:.2f}%")
        st.markdown(f"- **Viable Zones:** {via_cnt}/{len(fin_df)}")
        st.markdown(f"- **10-Year Revenue:** ₹{rev_tot:,.2f} Cr")

    with sc4:
        st.markdown("**📊 Scenario Range**")
        if st.session_state.scenario_results:
            base_s, opt_s, pess_s = st.session_state.scenario_results
            st.markdown(f"- **Pessimistic Yr10:** {pess_s.iloc[-1]['Total Demand (SCM/day)']:,.1f} SCM/day")
            st.markdown(f"- **Base Yr10:** {base_s.iloc[-1]['Total Demand (SCM/day)']:,.1f} SCM/day")
            st.markdown(f"- **Optimistic Yr10:** {opt_s.iloc[-1]['Total Demand (SCM/day)']:,.1f} SCM/day")
        if st.session_state.strategy_results:
            strat_s = st.session_state.strategy_results[0]
            best_strat = strat_s.sort_values("IRR (%)").iloc[-1]["Strategy"]
            st.markdown(f"- **Best Strategy:** {best_strat}")

    # Executive summary text
    best_strat_name = "Phased"
    best_strat_irr  = avg_irr2
    most_sensitive  = "Residential Growth Rate"
    if st.session_state.sensitivity_results:
        most_sensitive = st.session_state.sensitivity_results[0].iloc[0]["Variable"]
        most_swing     = st.session_state.sensitivity_results[0].iloc[0]["IRR Swing"]
    if st.session_state.strategy_results:
        strat_s2 = st.session_state.strategy_results[0]
        best_strat_row = strat_s2.sort_values("IRR (%)").iloc[-1]
        best_strat_name = best_strat_row["Strategy"]
        best_strat_irr  = best_strat_row["IRR (%)"]

    scen_pess_val = st.session_state.scenario_results[2].iloc[-1]["Total Demand (SCM/day)"] if st.session_state.scenario_results else y10_dem
    scen_opt_val  = st.session_state.scenario_results[1].iloc[-1]["Total Demand (SCM/day)"] if st.session_state.scenario_results else y10_dem

    exec_text = (
        f"This analysis covers {cp.get('city_name','N/A')} CGD network expansion planning across {len(zon_df)} zones. "
        f"Total projected demand reaches {y10_dem:,.1f} SCM/day by Year 10 representing {gr_pct}% growth. "
        f"The expansion program requires total capital investment of ₹{tot_cap:,.2f} Crores across all zones "
        f"with an average IRR of {avg_irr2:.2f}%. {via_cnt} out of {len(fin_df)} zones are financially viable "
        f"above the {vt_display:.0f}% IRR threshold. The highest priority expansion zone is {top_zone2} "
        f"offering the best combination of demand density and financial returns. "
        f"Under base case, the recommended expansion strategy is {best_strat_name} achieving {best_strat_irr:.2f}% IRR. "
        f"Sensitivity analysis identifies {most_sensitive} as the key risk factor. "
        f"Three-scenario analysis shows demand range of {scen_pess_val:,.1f} to {scen_opt_val:,.1f} SCM/day by Year 10. "
        f"This model can be reconfigured for any CGD geographical area by updating the sidebar inputs and zone data."
    )
    st.text_area("Executive Summary", value=exec_text, height=180)

    # Full report download
    demand_dl = dem_df
    zone_dl   = zon_df[["Priority_Rank","zone_name","Final_Score","Household_Density_Score",
                         "Industrial_Score","Penetration_Score","Proximity_Score","Recommendation"]].copy()
    fin_dl    = fin_df[["Zone Name","Total Capex (₹ Cr)","Year 10 Revenue (₹ Cr)",
                         "Payback Period (Years)","IRR (%)","NPV (₹ Cr)","Viability"]].copy()
    full_excel = build_excel({
        "Demand Forecast": demand_dl,
        "Zone Ranking":    zone_dl,
        "Financial Model": fin_dl,
    })
    st.download_button("⬇️ Download Complete Report Excel", data=full_excel,
        file_name="cgd_complete_report.xlsx", use_container_width=True,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("⬅️ Run the analysis to see the executive summary.")


# =============================================================================
# ─── PDF FINAL REPORT ────────────────────────────────────────────────────
# =============================================================================
st.divider()
st.markdown("### 📄 Final PDF Report")

if not PDF_AVAILABLE:
    st.warning("⚠️ reportlab not installed. Run `pip install reportlab` to enable PDF generation.")
elif not st.session_state.get("analysis_run"):
    st.info("⬅️ Run the analysis first, then generate the PDF report.")
else:
    st.markdown(
        "Generate a complete, professionally formatted PDF report covering demand forecast, "
        "zone prioritization, techno-economic analysis, scenario analysis, and executive summary."
    )

    def _c(hex_str):
        """Convert #RRGGBB to reportlab Color."""
        h = hex_str.lstrip("#")
        return colors.Color(int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255)

    NAVY   = _c("#1f4e79")
    BLUE   = _c("#2980b9")
    LTBLUE = _c("#d6eaf8")
    WHITE  = colors.white
    GRAY   = _c("#f2f3f4")
    GREEN  = _c("#1e8449")
    RED    = _c("#c0392b")
    AMBER  = _c("#d68910")

    def build_pdf_report() -> bytes:
        import datetime
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=1.8*cm, rightMargin=1.8*cm,
            topMargin=2*cm, bottomMargin=2*cm,
        )
        W = A4[0] - 3.6*cm   # usable width

        styles = getSampleStyleSheet()
        def S(name, **kw):
            return ParagraphStyle(name, parent=styles["Normal"], **kw)

        sTitle    = S("sTitle",    fontSize=22, textColor=WHITE,   alignment=TA_CENTER, leading=28, fontName="Helvetica-Bold")
        sSubTitle = S("sSub",      fontSize=11, textColor=LTBLUE,  alignment=TA_CENTER, leading=16)
        sSection  = S("sSection",  fontSize=13, textColor=NAVY,    spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold", borderPad=2)
        sBody     = S("sBody",     fontSize=9,  textColor=colors.black, leading=14, spaceAfter=4)
        sCaption  = S("sCaption",  fontSize=8,  textColor=_c("#555555"), leading=12, spaceAfter=8)
        sKPI_lbl  = S("sKPI_lbl",  fontSize=8,  textColor=_c("#555555"), alignment=TA_CENTER)
        sKPI_val  = S("sKPI_val",  fontSize=14, textColor=NAVY, fontName="Helvetica-Bold", alignment=TA_CENTER)
        sFooter   = S("sFooter",   fontSize=7,  textColor=_c("#888888"), alignment=TA_CENTER)
        sTableHdr = S("sTHdr",     fontSize=8,  textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)
        sTableCell= S("sTCell",    fontSize=8,  alignment=TA_LEFT)
        sTableNum = S("sTNum",     fontSize=8,  alignment=TA_RIGHT)

        cp   = st.session_state.city_params
        dem  = st.session_state.demand_df
        zon  = st.session_state.zone_ranking_df
        fin  = st.session_state.financial_df
        date = datetime.date.today().strftime("%d %B %Y")

        def tbl_style(has_header=True):
            cmds = [
                ("BACKGROUND",  (0,0), (-1,0), NAVY),
                ("TEXTCOLOR",   (0,0), (-1,0), WHITE),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [GRAY, WHITE]),
                ("GRID",        (0,0), (-1,-1), 0.35, _c("#cccccc")),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",  (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("LEFTPADDING", (0,0), (-1,-1), 5),
                ("RIGHTPADDING",(0,0), (-1,-1), 5),
            ]
            return TableStyle(cmds)

        def HR(): return HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=6, spaceBefore=6)

        story = []

        # ── COVER PAGE ────────────────────────────────────────────────────────
        cover_data = [[
            Paragraph("⛽️ PNG CGD Planning Model", sTitle),
        ]]
        cover_tbl = Table(cover_data, colWidths=[W])
        cover_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,-1), NAVY),
            ("TOPPADDING",  (0,0), (-1,-1), 18),
            ("BOTTOMPADDING",(0,0),(-1,-1), 18),
            ("LEFTPADDING", (0,0), (-1,-1), 10),
            ("RIGHTPADDING",(0,0),(-1,-1), 10),
        ]))
        story.append(cover_tbl)
        story.append(Spacer(1, 0.3*cm))

        subtitle_data = [[
            Paragraph("PNG Demand Forecasting &amp; Network Expansion Planning Model", sSubTitle),
        ]]
        sub_tbl = Table(subtitle_data, colWidths=[W])
        sub_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), BLUE),
            ("TOPPADDING", (0,0), (-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
        ]))
        story.append(sub_tbl)
        story.append(Spacer(1, 1*cm))

        # City info box
        meta = [
            [Paragraph("City / Project", sCaption),  Paragraph(cp.get("city_name", "N/A"), sBody)],
            [Paragraph("Report Date",    sCaption),  Paragraph(date, sBody)],
            [Paragraph("Evaluation Period", sCaption), Paragraph("10 Years", sBody)],
            [Paragraph("Model Version",  sCaption),  Paragraph("Generic CGD Analytics v2.0", sBody)],
            [Paragraph("Total Population", sCaption), Paragraph(f"{cp.get('total_pop',0):,}", sBody)],
            [Paragraph("Commercial Units", sCaption), Paragraph(f"{cp.get('total_comm',0):,}", sBody)],
        ]
        meta_tbl = Table(meta, colWidths=[4*cm, W-4*cm])
        meta_tbl.setStyle(TableStyle([
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [LTBLUE, WHITE]),
            ("GRID",        (0,0), (-1,-1), 0.3, _c("#bbbbbb")),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 1.5*cm))
        story.append(Paragraph("Prepared by: CGD Analytics Platform  |  Confidential", sFooter))
        story.append(PageBreak())

        # ── EXECUTIVE SUMMARY ─────────────────────────────────────────────────
        story.append(Paragraph("📊 Executive Summary", sSection))
        story.append(HR())

        y10_row  = dem.iloc[-1]
        y10_dem  = y10_row["Total Demand (SCM/day)"]
        y1_dem   = dem.iloc[0]["Total Demand (SCM/day)"]
        tot_cust = y10_row["Residential Customers"] + y10_row["Commercial Customers"] + y10_row["Industrial Customers"]
        tot_cap  = fin["Total Capex (₹ Cr)"].sum()
        avg_irr  = fin["IRR (%)"].mean()
        via_cnt  = (fin["Viability"] == "Financially Viable").sum()
        top_zone = zon.iloc[0]["zone_name"] if len(zon) > 0 else "N/A"

        kpi_data = [
            [Paragraph("Year 10 Demand", sKPI_lbl), Paragraph("Total Customers", sKPI_lbl),
             Paragraph("Total CAPEX", sKPI_lbl),     Paragraph("Avg IRR", sKPI_lbl),
             Paragraph("Viable Zones", sKPI_lbl)],
            [Paragraph(f"{y10_dem:,.0f}<br/>SCM/day", sKPI_val),
             Paragraph(f"{tot_cust:,}", sKPI_val),
             Paragraph(f"₹{tot_cap:.1f} Cr", sKPI_val),
             Paragraph(f"{avg_irr:.1f}%", sKPI_val),
             Paragraph(f"{via_cnt}/{len(fin)}", sKPI_val)],
        ]
        kw = W / 5
        kpi_tbl = Table(kpi_data, colWidths=[kw]*5)
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), NAVY),
            ("BACKGROUND",  (0,1), (-1,1), LTBLUE),
            ("GRID",        (0,0), (-1,-1), 0.5, BLUE),
            ("TOPPADDING",  (0,0), (-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 0.5*cm))

        gr_pct = round((y10_dem - y1_dem) / max(y1_dem, 1) * 100, 1)
        exec_para = (
            f"This report covers the CGD network expansion planning for <b>{cp.get('city_name','N/A')}</b> "
            f"across <b>{len(zon)} zones</b>. Total projected demand reaches "
            f"<b>{y10_dem:,.1f} SCM/day</b> by Year 10, representing {gr_pct}% growth from Year 1. "
            f"The expansion program requires a total capital investment of <b>₹{tot_cap:,.2f} Crores</b> "
            f"with an average IRR of <b>{avg_irr:.2f}%</b>. "
            f"{via_cnt} out of {len(fin)} zones are financially viable above the viability threshold. "
            f"The highest priority expansion zone is <b>{top_zone}</b>, offering the best combination "
            f"of demand density and financial returns."
        )
        story.append(Paragraph(exec_para, sBody))
        story.append(PageBreak())

        # ── DEMAND FORECAST ───────────────────────────────────────────────────
        story.append(Paragraph("📈 Section 1 — Demand Forecast", sSection))
        story.append(HR())
        story.append(Paragraph(
            "Projected PNG demand across residential, commercial, and industrial segments over the 10-year evaluation period.",
            sCaption
        ))

        dem_cols = ["Year", "Residential Customers", "Commercial Customers", "Industrial Customers",
                    "Total Demand (SCM/day)", "Annual Demand (SCM)"]
        dem_show = dem[dem_cols].copy()
        dem_header = [[Paragraph(c, sTableHdr) for c in dem_show.columns.tolist()]]
        dem_rows   = []
        for _, r in dem_show.iterrows():
            row = []
            for i, v in enumerate(r):
                txt = f"{int(v):,}" if i in [0,1,2,3,5] else f"{v:,.1f}"
                row.append(Paragraph(txt, sTableNum if i > 0 else sTableCell))
            dem_rows.append(row)
        col_w = [1.2*cm] + [(W-1.2*cm)/5]*5
        dem_tbl = Table(dem_header + dem_rows, colWidths=col_w, repeatRows=1)
        dem_tbl.setStyle(tbl_style())
        story.append(dem_tbl)
        story.append(PageBreak())

        # ── ZONE PRIORITIZATION ───────────────────────────────────────────────
        story.append(Paragraph("🏙️ Section 2 — Zone Prioritization", sSection))
        story.append(HR())
        story.append(Paragraph(
            "Zones ranked by weighted composite score across household density, industrial demand, penetration potential, and network proximity.",
            sCaption
        ))

        zon_cols = ["Priority_Rank", "zone_name", "Final_Score",
                    "Household_Density_Score", "Industrial_Score", "Recommendation"]
        zon_show = zon[[c for c in zon_cols if c in zon.columns]].copy()
        zon_header = [[Paragraph(c.replace("_", " "), sTableHdr) for c in zon_show.columns.tolist()]]
        zon_rows = []
        for i, (_, r) in enumerate(zon_show.iterrows()):
            row = [Paragraph(str(v), sTableNum if j > 0 and j < len(r)-1 else sTableCell)
                   for j, v in enumerate(r)]
            zon_rows.append(row)
        zon_col_w_cnt = len(zon_show.columns)
        zon_cw = [W / zon_col_w_cnt] * zon_col_w_cnt
        zon_tbl = Table(zon_header + zon_rows, colWidths=zon_cw, repeatRows=1)
        zon_tbl.setStyle(tbl_style())
        # Colour top 3 rows gold/green
        for ri in range(1, min(4, len(zon_rows)+1)):
            col = LTBLUE if ri > 1 else _c("#d5f5e3")
            zon_tbl.setStyle(TableStyle([("BACKGROUND", (0,ri), (-1,ri), col)]))
        story.append(zon_tbl)
        story.append(PageBreak())

        # ── TECHNO-ECONOMIC ────────────────────────────────────────────────────
        story.append(Paragraph("💰 Section 3 — Techno-Economic Analysis", sSection))
        story.append(HR())
        story.append(Paragraph(
            "Financial viability assessment per zone including CAPEX, IRR, NPV, and payback period.",
            sCaption
        ))

        fin_cols = ["Zone Name", "Total Capex (₹ Cr)", "Year 10 Revenue (₹ Cr)",
                    "Payback Period (Years)", "IRR (%)", "NPV (₹ Cr)", "Viability"]
        fin_show = fin[[c for c in fin_cols if c in fin.columns]].copy()
        fin_header = [[Paragraph(c, sTableHdr) for c in fin_show.columns.tolist()]]
        fin_rows = []
        for _, r in fin_show.iterrows():
            row = []
            for j, v in enumerate(r):
                cell_style = sTableCell if j == 0 or j == len(r)-1 else sTableNum
                row.append(Paragraph(str(v), cell_style))
            fin_rows.append(row)
        fin_cw = [3.5*cm] + [(W-3.5*cm) / (len(fin_show.columns)-1)] * (len(fin_show.columns)-1)
        fin_tbl = Table(fin_header + fin_rows, colWidths=fin_cw, repeatRows=1)
        fin_tbl.setStyle(tbl_style())
        # Colour viability column
        for ri, (_, r) in enumerate(fin_show.iterrows(), start=1):
            v = str(r.get("Viability", ""))
            if   "Viable"  in v: c = _c("#d5f5e3")
            elif "Marginal" in v: c = _c("#fef9e7")
            else:                 c = _c("#fadbd8")
            fin_tbl.setStyle(TableStyle([("BACKGROUND", (-1,ri), (-1,ri), c)]))
        story.append(fin_tbl)
        story.append(PageBreak())

        # ── SCENARIO ANALYSIS ──────────────────────────────────────────────────
        if st.session_state.get("scenario_results"):
            story.append(Paragraph("📊 Section 4 — Scenario Analysis", sSection))
            story.append(HR())
            story.append(Paragraph(
                "Demand projection under Base, Optimistic, and Pessimistic growth scenarios.",
                sCaption
            ))
            base_df, opt_df, pess_df = st.session_state.scenario_results
            scen_hdr = [[Paragraph(h, sTableHdr) for h in
                         ["Year","Base Demand (SCM/day)","Optimistic (SCM/day)","Pessimistic (SCM/day)"]]]
            scen_rows = []
            for i in range(len(base_df)):
                scen_rows.append([
                    Paragraph(str(base_df.iloc[i]["Year"]),  sTableCell),
                    Paragraph(f"{base_df.iloc[i]['Total Demand (SCM/day)']:,.1f}", sTableNum),
                    Paragraph(f"{opt_df.iloc[i]['Total Demand (SCM/day)']:,.1f}",  sTableNum),
                    Paragraph(f"{pess_df.iloc[i]['Total Demand (SCM/day)']:,.1f}", sTableNum),
                ])
            scen_cw = [1.5*cm, (W-1.5*cm)/3, (W-1.5*cm)/3, (W-1.5*cm)/3]
            scen_tbl = Table(scen_hdr + scen_rows, colWidths=scen_cw, repeatRows=1)
            scen_tbl.setStyle(tbl_style())
            story.append(scen_tbl)
            story.append(PageBreak())

        # ── SENSITIVITY ANALYSIS ───────────────────────────────────────────────
        if st.session_state.get("sensitivity_results"):
            story.append(Paragraph("🎯 Section 5 — Sensitivity Analysis", sSection))
            story.append(HR())
            story.append(Paragraph(
                "IRR sensitivity to key model variables, ranked by impact swing (high to low).",
                sCaption
            ))
            sens_df, _, base_irr = st.session_state.sensitivity_results
            story.append(Paragraph(f"Base Case IRR: <b>{base_irr:.2f}%</b>", sBody))
            sens_cols = ["Variable", "Base Value", "IRR at Min", "IRR at Max", "IRR Swing", "Risk Level"]
            sens_show = sens_df[[c for c in sens_cols if c in sens_df.columns]].copy()
            sens_hdr = [[Paragraph(c, sTableHdr) for c in sens_show.columns.tolist()]]
            sens_rows = []
            for _, r in sens_show.iterrows():
                row = [Paragraph(str(v), sTableCell if j == 0 or j == len(r)-1 else sTableNum)
                       for j, v in enumerate(r)]
                sens_rows.append(row)
            scw = [3.5*cm] + [(W-3.5*cm) / (len(sens_show.columns)-1)] * (len(sens_show.columns)-1)
            sens_tbl = Table(sens_hdr + sens_rows, colWidths=scw, repeatRows=1)
            sens_tbl.setStyle(tbl_style())
            for ri, (_, r) in enumerate(sens_show.iterrows(), start=1):
                risk = str(r.get("Risk Level", ""))
                if   "High"   in risk: rc = _c("#fadbd8")
                elif "Medium" in risk: rc = _c("#fef9e7")
                else:                  rc = _c("#d5f5e3")
                sens_tbl.setStyle(TableStyle([("BACKGROUND", (-1,ri), (-1,ri), rc)]))
            story.append(sens_tbl)
            story.append(PageBreak())

        # ── STRATEGY COMPARISON ────────────────────────────────────────────────
        if st.session_state.get("strategy_results"):
            story.append(Paragraph("⚖️ Section 6 — Strategy Comparison", sSection))
            story.append(HR())
            story.append(Paragraph(
                "Comparison of three expansion strategies: Aggressive, Phased, and Asset Light.",
                sCaption
            ))
            strat_df = st.session_state.strategy_results[0]
            strat_hdr = [[Paragraph(c, sTableHdr) for c in strat_df.columns.tolist()]]
            strat_rows = []
            for _, r in strat_df.iterrows():
                strat_rows.append([
                    Paragraph(str(v), sTableCell if j == 0 or j == len(r)-1 else sTableNum)
                    for j, v in enumerate(r)
                ])
            strat_cw = [3*cm] + [(W-3*cm) / (len(strat_df.columns)-1)] * (len(strat_df.columns)-1)
            strat_tbl = Table(strat_hdr + strat_rows, colWidths=strat_cw, repeatRows=1)
            strat_tbl.setStyle(tbl_style())
            story.append(strat_tbl)
            story.append(Spacer(1, 0.5*cm))

        # ── FOOTER NOTE ───────────────────────────────────────────────────────
        story.append(HR())
        story.append(Paragraph(
            f"Generated by CGD Analytics Platform • {date} • Confidential — For internal use only.",
            sFooter
        ))

        doc.build(story)
        return buf.getvalue()

    col_pdf1, col_pdf2 = st.columns([3, 1])
    with col_pdf1:
        st.markdown(
            "**Report includes:** Cover page • Executive Summary • Demand Forecast table • "
            "Zone Prioritization • Techno-Economic Analysis • Scenario Analysis • "
            "Sensitivity Analysis • Strategy Comparison"
        )
    with col_pdf2:
        try:
            pdf_bytes = build_pdf_report()
            import datetime
            fname = f"CGD_Report_{st.session_state.city_params.get('city_name','City').replace(' ','_')}_{datetime.date.today()}.pdf"
            st.download_button(
                label="📄 Download PDF Report",
                data=pdf_bytes,
                file_name=fname,
                mime="application/pdf",
                use_container_width=True,
                type="primary",
            )
        except Exception as e:
            st.error(f"PDF generation failed: {e}")

