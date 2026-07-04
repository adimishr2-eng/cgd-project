"""
demand_forecast.py
==================

PNG (Piped Natural Gas) Demand Forecasting Module for City Gas Distribution (CGD) Networks.

This module reads city-level demographic and consumption parameters from a CSV file,
projects demand over a 10-year horizon across three customer segments (residential,
commercial, industrial), exports the results to a formatted Excel workbook, generates
publication-quality charts, and prints a concise summary to the console.

Key assumptions
---------------
- Penetration rates grow geometrically from a base value and are capped at 85 %.
- Daily demand per connected customer is constant across years (no per-capita growth).
- A year is taken as 365 days for annual conversion.

Author : CGD Analytics Team
Created: 2026-06-27
"""

import os
import math

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
PENETRATION_CAP = 0.85          # Maximum penetration rate (85 %)
PROJECTION_YEARS = 10           # Forecast horizon in years
DAYS_PER_YEAR = 365             # Days used for annual conversion
CHART_DPI = 300                 # Resolution for saved chart images
MATPLOTLIB_STYLE = "ggplot"     # Consistent professional styling


# ---------------------------------------------------------------------------
# Helper: parse city_data.csv into a typed dictionary
# ---------------------------------------------------------------------------
def _parse_city_data(csv_path: str) -> dict:
    """Read the parameter / value CSV and return a dictionary with typed values.

    Parameters
    ----------
    csv_path : str
        Absolute path to ``city_data.csv``.  The file must contain two columns:
        ``parameter`` (str) and ``value`` (str representation of a number or text).

    Returns
    -------
    dict
        Keys are parameter names (str), values are converted to ``float`` where
        possible, otherwise kept as ``str``.

    Raises
    ------
    FileNotFoundError
        If *csv_path* does not exist.
    """
    df = pd.read_csv(csv_path)

    # Strip any accidental whitespace in column names and values
    df.columns = df.columns.str.strip()
    df["parameter"] = df["parameter"].astype(str).str.strip()
    df["value"] = df["value"].astype(str).str.strip()

    city_dict = {}
    for _, row in df.iterrows():
        key = row["parameter"]
        raw = row["value"]
        # Attempt numeric conversion; fall back to string
        try:
            city_dict[key] = float(raw)
        except ValueError:
            city_dict[key] = raw

    return city_dict


# ---------------------------------------------------------------------------
# Helper: compute capped penetration for a given year
# ---------------------------------------------------------------------------
def _penetration(base_rate: float, growth_rate: float, year: int) -> float:
    """Return the penetration rate for *year*, capped at PENETRATION_CAP.

    The penetration grows geometrically:
        penetration = base_rate * (1 + growth_rate) ^ (year - 1)

    Parameters
    ----------
    base_rate : float
        Year-1 penetration fraction (e.g. 0.10 for 10 %).
    growth_rate : float
        Annual growth rate as a fraction (e.g. 0.08 for 8 %).
    year : int
        Projection year (1-indexed).

    Returns
    -------
    float
        Penetration fraction, at most ``PENETRATION_CAP``.
    """
    return min(base_rate * (1 + growth_rate) ** (year - 1), PENETRATION_CAP)


# ---------------------------------------------------------------------------
# Core: build the 10-year demand projection DataFrame
# ---------------------------------------------------------------------------
def _build_projection(city: dict) -> pd.DataFrame:
    """Create a year-by-year demand projection DataFrame.

    Parameters
    ----------
    city : dict
        Parsed city data dictionary (output of ``_parse_city_data``).

    Returns
    -------
    pd.DataFrame
        Columns defined by the specification (Year, customers, demands, totals).
    """
    # ---- Derive total households from population and average household size ----
    total_households = round(city["total_population"] / city["avg_household_size"])

    # ---- Retrieve base penetration rates and growth rates ----
    res_base = city["base_residential_penetration"]
    res_growth = city["residential_growth_rate"]
    res_daily = city["residential_consumption_scm_per_day"]

    com_base = city["base_commercial_penetration"]
    com_growth = city["commercial_growth_rate"]
    com_daily = city["commercial_consumption_scm_per_day"]

    ind_base = city["base_industrial_penetration"]
    ind_growth = city["industrial_growth_rate"]
    ind_daily = city["industrial_consumption_scm_per_day"]

    total_commercial = city["total_commercial_establishments"]
    total_industrial = city["total_industrial_units"]

    # ---- Year-by-year projection ----
    records = []
    for year in range(1, PROJECTION_YEARS + 1):
        # --- Residential ---
        res_pen = _penetration(res_base, res_growth, year)
        res_customers = round(total_households * res_pen)
        res_demand = res_customers * res_daily

        # --- Commercial ---
        com_pen = _penetration(com_base, com_growth, year)
        com_customers = round(total_commercial * com_pen)
        com_demand = com_customers * com_daily

        # --- Industrial ---
        ind_pen = _penetration(ind_base, ind_growth, year)
        ind_customers = round(total_industrial * ind_pen)
        ind_demand = ind_customers * ind_daily

        # --- Totals ---
        total_daily = res_demand + com_demand + ind_demand
        annual_demand = total_daily * DAYS_PER_YEAR

        records.append(
            {
                "Year": year,
                "Residential_Customers": res_customers,
                "Commercial_Customers": com_customers,
                "Industrial_Customers": ind_customers,
                "Residential_Demand_SCM_per_day": res_demand,
                "Commercial_Demand_SCM_per_day": com_demand,
                "Industrial_Demand_SCM_per_day": ind_demand,
                "Total_Demand_SCM_per_day": total_daily,
                "Annual_Demand_SCM": annual_demand,
            }
        )

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Export: formatted Excel workbook
# ---------------------------------------------------------------------------
def _export_to_excel(df: pd.DataFrame, excel_path: str) -> None:
    """Write *df* to an Excel file with professional formatting.

    Formatting includes bold headers, auto-adjusted column widths, and
    comma-separated number format for large numeric columns.

    Parameters
    ----------
    df : pd.DataFrame
        The demand forecast DataFrame.
    excel_path : str
        Destination ``.xlsx`` file path.
    """
    # Write raw data first, then dress it up with openpyxl
    df.to_excel(excel_path, index=False, sheet_name="Demand Forecast", engine="openpyxl")

    wb = load_workbook(excel_path)
    ws = wb.active

    # ---- Bold header row ----
    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ---- Number formatting with commas for all data cells ----
    comma_format = "#,##0"          # integers with thousand separators
    comma_float = "#,##0.00"        # two-decimal floats (demand values)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Use integer format for customer counts and Year column,
                # float format for demand values
                if cell.column <= 4:        # Year + 3 customer columns
                    cell.number_format = comma_format
                else:
                    cell.number_format = comma_float
            cell.alignment = Alignment(horizontal="center")

    # ---- Auto-adjust column widths ----
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        # Add a little padding
        ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(excel_path)


# ---------------------------------------------------------------------------
# Chart 1: Stacked bar chart of daily demand by segment
# ---------------------------------------------------------------------------
def _chart_stacked_bar(df: pd.DataFrame, save_path: str) -> None:
    """Generate a stacked bar chart of daily PNG demand by customer segment.

    Parameters
    ----------
    df : pd.DataFrame
        Demand forecast DataFrame.
    save_path : str
        File path for the saved PNG image.
    """
    plt.style.use(MATPLOTLIB_STYLE)
    fig, ax = plt.subplots(figsize=(14, 8))

    years = df["Year"].values
    res = df["Residential_Demand_SCM_per_day"].values
    com = df["Commercial_Demand_SCM_per_day"].values
    ind = df["Industrial_Demand_SCM_per_day"].values
    totals = df["Total_Demand_SCM_per_day"].values

    x = np.arange(len(years))
    bar_width = 0.55

    # Stack: residential at bottom, commercial in the middle, industrial on top
    ax.bar(x, res, width=bar_width, label="Residential", color="#4A90D9")
    ax.bar(x, com, width=bar_width, bottom=res, label="Commercial", color="#F5A623")
    ax.bar(x, ind, width=bar_width, bottom=res + com, label="Industrial", color="#7ED321")

    # Value labels showing total on top of each stacked bar
    for i, total in enumerate(totals):
        ax.text(
            x[i],
            total + totals.max() * 0.01,  # slight offset above bar
            f"{total:,.0f}",
            ha="center",
            va="bottom",
            fontsize=8,
            fontweight="bold",
        )

    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Total Demand (SCM / day)", fontsize=12)
    ax.set_title(
        "PNG Demand Forecast by Customer Segment (Year 1-10)",
        fontsize=14,
        fontweight="bold",
    )
    ax.set_xticks(x)
    ax.set_xticklabels([f"Year {y}" for y in years])
    ax.legend(loc="upper left", fontsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.7)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))

    fig.tight_layout()
    fig.savefig(save_path, dpi=CHART_DPI, bbox_inches="tight")
    plt.close("all")  # Free memory


# ---------------------------------------------------------------------------
# Chart 2: Line chart of annual demand growth curve
# ---------------------------------------------------------------------------
def _chart_growth_curve(df: pd.DataFrame, save_path: str) -> None:
    """Generate a line chart with shaded area showing annual demand growth.

    Parameters
    ----------
    df : pd.DataFrame
        Demand forecast DataFrame.
    save_path : str
        File path for the saved PNG image.
    """
    plt.style.use(MATPLOTLIB_STYLE)
    fig, ax = plt.subplots(figsize=(12, 7))

    years = df["Year"].values
    annual = df["Annual_Demand_SCM"].values

    # Line with circle markers
    ax.plot(
        years,
        annual,
        marker="o",
        markersize=8,
        linewidth=2.5,
        color="#4A90D9",
        label="Annual Demand (SCM)",
    )

    # Shade the area under the curve
    ax.fill_between(years, annual, alpha=0.3, color="#4A90D9")

    # Data labels at each point
    for yr, val in zip(years, annual):
        ax.annotate(
            f"{val:,.0f}",
            xy=(yr, val),
            xytext=(0, 12),
            textcoords="offset points",
            ha="center",
            fontsize=8,
            fontweight="bold",
        )

    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Annual Demand (SCM)", fontsize=12)
    ax.set_title(
        "Total PNG Demand Growth Curve (10-Year Projection)",
        fontsize=14,
        fontweight="bold",
    )
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.grid(axis="both", linestyle="--", alpha=0.7)
    ax.legend(loc="upper left", fontsize=10)

    fig.tight_layout()
    fig.savefig(save_path, dpi=CHART_DPI, bbox_inches="tight")
    plt.close("all")  # Free memory


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------
def _print_summary(city: dict, df: pd.DataFrame) -> None:
    """Print a human-readable summary of key forecast metrics.

    Parameters
    ----------
    city : dict
        Parsed city data dictionary.
    df : pd.DataFrame
        Demand forecast DataFrame.
    """
    total_households = round(city["total_population"] / city["avg_household_size"])
    y1 = df.loc[df["Year"] == 1, "Total_Demand_SCM_per_day"].values[0]
    y5 = df.loc[df["Year"] == 5, "Total_Demand_SCM_per_day"].values[0]
    y10 = df.loc[df["Year"] == 10, "Total_Demand_SCM_per_day"].values[0]
    pct_growth = ((y10 - y1) / y1) * 100 if y1 != 0 else float("inf")

    print("\n" + "=" * 60)
    print("  PNG DEMAND FORECAST -- SUMMARY")
    print("=" * 60)
    print(f"  City Name              : {city.get('city_name', 'N/A')}")
    print(f"  Total Households       : {total_households:,}")
    print(f"  Year 1 Total Demand    : {y1:,.2f} SCM/day")
    print(f"  Year 5 Total Demand    : {y5:,.2f} SCM/day")
    print(f"  Year 10 Total Demand   : {y10:,.2f} SCM/day")
    print(f"  Growth (Yr 1 -> Yr 10)  : {pct_growth:,.1f} %")
    print("=" * 60 + "\n")


# ---------------------------------------------------------------------------
# Main orchestration function
# ---------------------------------------------------------------------------
def run_demand_forecast(base_path: str) -> pd.DataFrame:
    """Execute the full PNG demand forecasting pipeline.

    The pipeline performs the following steps:
    1. Read and parse ``inputs/city_data.csv``.
    2. Calculate 10-year demand projections for residential, commercial,
       and industrial segments.
    3. Export a formatted Excel workbook to ``outputs/demand_forecast.xlsx``.
    4. Generate two publication-quality charts under ``outputs/charts/``.
    5. Print a concise forecast summary to the console.

    Parameters
    ----------
    base_path : str
        Absolute path to the project root directory (e.g.
        ``'C:/path/to/cgd_project'``).  All relative paths are resolved
        from this root using ``os.path.join``.

    Returns
    -------
    pd.DataFrame
        The 10-year demand forecast table with columns for year, customer
        counts, daily demands, and annual demand.
    """
    # ------------------------------------------------------------------ #
    # Step 1: Read city data from CSV                                     #
    # ------------------------------------------------------------------ #
    print("Reading city data...")
    csv_path = os.path.join(base_path, "inputs", "city_data.csv")
    city = _parse_city_data(csv_path)
    print(f"  -> Loaded {len(city)} parameters for '{city.get('city_name', 'unknown city')}'.")

    # ------------------------------------------------------------------ #
    # Step 2: Calculate demand projections                                #
    # ------------------------------------------------------------------ #
    print("Calculating demand projections...")
    df = _build_projection(city)
    print(f"  -> Projected demand for {PROJECTION_YEARS} years.")

    # ------------------------------------------------------------------ #
    # Step 3: Export formatted Excel workbook                             #
    # ------------------------------------------------------------------ #
    print("Exporting to Excel...")
    outputs_dir = os.path.join(base_path, "outputs")
    os.makedirs(outputs_dir, exist_ok=True)
    excel_path = os.path.join(outputs_dir, "demand_forecast.xlsx")
    _export_to_excel(df, excel_path)
    print(f"  -> Saved workbook to {excel_path}")

    # ------------------------------------------------------------------ #
    # Step 4: Generate charts                                             #
    # ------------------------------------------------------------------ #
    print("Generating charts...")
    charts_dir = os.path.join(outputs_dir, "charts")
    os.makedirs(charts_dir, exist_ok=True)

    # Chart 1 -- Stacked bar chart
    bar_path = os.path.join(charts_dir, "demand_forecast_stacked_bar.png")
    _chart_stacked_bar(df, bar_path)
    print(f"  -> Stacked bar chart saved to {bar_path}")

    # Chart 2 -- Growth curve line chart
    line_path = os.path.join(charts_dir, "demand_growth_curve.png")
    _chart_growth_curve(df, line_path)
    print(f"  -> Growth curve chart saved to {line_path}")

    # ------------------------------------------------------------------ #
    # Step 5: Console summary                                             #
    # ------------------------------------------------------------------ #
    _print_summary(city, df)

    print("Demand forecasting complete.\n")
    return df


# ---------------------------------------------------------------------------
# Entry point for standalone execution
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # When run directly, assume the project root is one level above the
    # directory containing this script (i.e. cgd_project/).
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)

    run_demand_forecast(project_root)
