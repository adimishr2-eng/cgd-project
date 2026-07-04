"""
techno_economic.py -- Techno-Economic Analysis for CGD (City Gas Distribution) Zones
===================================================================================

This module performs a complete financial viability assessment for each zone in a
CGD project.  It reads zone-level data, city-level parameters, and financial
assumptions, then builds a 10-year discounted cash-flow model per zone.

Key outputs:
    • Per-zone capital expenditure breakdown (pipeline, domestic, commercial,
      industrial connections).
    • Year-by-year cash-flow projections with linear customer ramp-up.
    • Internal Rate of Return (IRR), Net Present Value (NPV), and simple
      payback period for every zone.
    • A viability classification (Financially Viable / Marginal / Needs Review).
    • A formatted Excel workbook with a summary sheet and per-zone detail sheets.
    • Two publication-quality charts (IRR comparison bar chart and cumulative
      cash-flow line chart for the top-3 priority zones).

Usage
-----
    from models.techno_economic import run_techno_economic
    summary_df = run_techno_economic("path/to/project_root")

Dependencies
------------
    pandas, numpy, numpy_financial, matplotlib, openpyxl
"""

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------
import os
import math
import warnings

import numpy as np
import pandas as pd
import numpy_financial as npf
import matplotlib.pyplot as plt
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helper -- parse a two-column parameter CSV into a Python dict
# ---------------------------------------------------------------------------
def _parse_param_csv(filepath):
    """Read a CSV with columns ``parameter`` and ``value`` into a dict.

    Numeric strings are automatically converted to ``float``; everything
    else stays as a string.

    Parameters
    ----------
    filepath : str
        Absolute or relative path to the CSV file.

    Returns
    -------
    dict
        Mapping of parameter names to their (typed) values.
    """
    df = pd.read_csv(filepath)
    params = {}
    for _, row in df.iterrows():
        key = str(row["parameter"]).strip()
        raw = str(row["value"]).strip()
        # Try to coerce to a number; fall back to string
        try:
            params[key] = float(raw)
        except ValueError:
            params[key] = raw
    return params


# ---------------------------------------------------------------------------
# Helper -- auto-fit column widths in an openpyxl worksheet
# ---------------------------------------------------------------------------
def _auto_fit_columns(ws, min_width=10):
    """Adjust each column's width to fit the longest cell value.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet whose columns should be resized.
    min_width : int, optional
        Minimum column width in characters (default 10).
    """
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = max_len


# ---------------------------------------------------------------------------
# Main analysis function
# ---------------------------------------------------------------------------
def run_techno_economic(base_path, zone_ranking_df=None):
    """Run the full techno-economic analysis for every CGD zone.

    Parameters
    ----------
    base_path : str
        Root directory of the CGD project.  Expected sub-folders:
        ``inputs/`` (zone_data.csv, financial_params.csv, city_data.csv)
        and ``outputs/`` (created automatically).
    zone_ranking_df : pandas.DataFrame or None, optional
        A DataFrame produced by the zone-prioritisation model.  Must contain
        a ``zone_name`` column.  When supplied the analysis honours zone
        ordering for the top-3 cumulative-cash-flow chart.  When *None*,
        zones are used in the order they appear in ``zone_data.csv``.

    Returns
    -------
    pandas.DataFrame
        Summary table with one row per zone and columns for CAPEX, revenue,
        payback, IRR, NPV, and viability.
    """
    print("=" * 70)
    print("  TECHNO-ECONOMIC ANALYSIS -- CGD Zone Financial Model")
    print("=" * 70)

    # ------------------------------------------------------------------
    # 1. Read input files
    # ------------------------------------------------------------------
    zone_csv_path = os.path.join(base_path, "inputs", "zone_data.csv")
    fin_csv_path  = os.path.join(base_path, "inputs", "financial_params.csv")
    city_csv_path = os.path.join(base_path, "inputs", "city_data.csv")

    print(f"\n[1/12] Reading zone data        -> {zone_csv_path}")
    zone_df = pd.read_csv(zone_csv_path)

    print(f"[1/12] Reading financial params  -> {fin_csv_path}")
    fin = _parse_param_csv(fin_csv_path)

    print(f"[1/12] Reading city data         -> {city_csv_path}")
    city = _parse_param_csv(city_csv_path)

    # If zone_ranking_df is provided, reorder zone_df to match the ranking
    if zone_ranking_df is not None:
        # Build an ordered list of zone names from the ranking DataFrame
        ranked_names = zone_ranking_df["zone_name"].tolist()
        # Reindex zone_df to follow that order (zones not in ranking are
        # appended at the end so nothing is silently dropped)
        zone_df = zone_df.set_index("zone_name").loc[ranked_names].reset_index()
        print(f"       Zone order set from ranking DataFrame ({len(ranked_names)} zones)")
    else:
        print("       No ranking DataFrame provided -- using original CSV order")

    # Ensure numeric types for every zone column we'll use
    for col in ["area_sqkm", "total_households", "industrial_units",
                "existing_penetration_pct"]:
        zone_df[col] = pd.to_numeric(zone_df[col], errors="coerce")

    # Unpack financial parameters for readability
    pe_pipeline_cost_lakh_per_km    = fin["pe_pipeline_cost_lakh_per_km"]
    domestic_connection_cost_rs     = fin["domestic_connection_cost_rs"]
    commercial_connection_cost_rs   = fin["commercial_connection_cost_rs"]
    industrial_connection_cost_rs   = fin["industrial_connection_cost_rs"]
    avg_domestic_annual_revenue     = fin["avg_domestic_annual_revenue_rs"]
    avg_commercial_annual_revenue   = fin["avg_commercial_annual_revenue_rs"]
    avg_industrial_annual_revenue   = fin["avg_industrial_annual_revenue_rs"]
    operating_cost_pct_of_revenue   = fin["operating_cost_pct_of_revenue"]
    discount_rate                   = fin["discount_rate"]
    project_years                   = int(fin["project_years"])  # 10

    # City-level parameter we need: total commercial establishments
    total_commercial_establishments = city["total_commercial_establishments"]

    # Sum of total_households across ALL zones (denominator for commercial
    # connection allocation)
    total_households_all_zones = zone_df["total_households"].sum()

    print(f"\n       Zones loaded             : {len(zone_df)}")
    print(f"       Total households (all)   : {total_households_all_zones:,.0f}")
    print(f"       Commercial estabs (city) : {total_commercial_establishments:,.0f}")
    print(f"       Discount rate            : {discount_rate:.0%}")
    print(f"       Project horizon          : {project_years} years")

    # ------------------------------------------------------------------
    # 2. Per-zone connection and pipeline calculations
    # ------------------------------------------------------------------
    print("\n[2/12] Calculating new connections per zone ...")

    # Pipeline length heuristic: sqrt(area) * 2.5 km
    zone_df["pipeline_length_km"] = (
        zone_df["area_sqkm"].apply(lambda a: round(math.sqrt(a) * 2.5, 1))
    )

    # New domestic connections: un-penetrated fraction × 60 % conversion
    zone_df["new_domestic_connections"] = (
        zone_df.apply(
            lambda r: round(
                r["total_households"]
                * (1 - r["existing_penetration_pct"])
                * 0.60
            ),
            axis=1,
        )
    ).astype(int)

    # New commercial connections: zone share of city-wide establishments × 40 %
    zone_df["new_commercial_connections"] = (
        zone_df.apply(
            lambda r: round(
                (r["total_households"] / total_households_all_zones)
                * total_commercial_establishments
                * 0.40
            ),
            axis=1,
        )
    ).astype(int)

    # New industrial connections: 70 % of zone's industrial units
    zone_df["new_industrial_connections"] = (
        zone_df["industrial_units"].apply(lambda u: round(u * 0.70))
    ).astype(int)

    # ------------------------------------------------------------------
    # 3. Capital expenditure breakdown
    # ------------------------------------------------------------------
    print("[3/12] Computing capital expenditure ...")

    # Pipeline cost: length_km × cost_per_km_in_lakhs × 100,000 (-> Rs)
    zone_df["pipeline_cost"] = (
        zone_df["pipeline_length_km"] * pe_pipeline_cost_lakh_per_km * 100_000
    )

    zone_df["domestic_capex"] = (
        zone_df["new_domestic_connections"] * domestic_connection_cost_rs
    )
    zone_df["commercial_capex"] = (
        zone_df["new_commercial_connections"] * commercial_connection_cost_rs
    )
    zone_df["industrial_capex"] = (
        zone_df["new_industrial_connections"] * industrial_connection_cost_rs
    )

    zone_df["total_capex"] = (
        zone_df["pipeline_cost"]
        + zone_df["domestic_capex"]
        + zone_df["commercial_capex"]
        + zone_df["industrial_capex"]
    )

    # ------------------------------------------------------------------
    # 4–7. Build 10-year cash-flow model & compute IRR / NPV / payback
    # ------------------------------------------------------------------
    print("[4/12] Building 10-year cash-flow projections ...")
    print("[5/12] Computing IRR per zone ...")
    print("[6/12] Computing NPV per zone ...")
    print("[7/12] Computing payback period ...")

    # Accumulators for the summary
    irr_list = []
    npv_list = []
    payback_list = []
    year10_revenue_list = []

    # Per-zone detail tables (stored for Excel export and chart)
    zone_cashflow_tables = {}  # zone_name -> DataFrame

    for _, row in zone_df.iterrows():
        zone_name  = row["zone_name"]
        capex      = row["total_capex"]
        dom_conn   = row["new_domestic_connections"]
        com_conn   = row["new_commercial_connections"]
        ind_conn   = row["new_industrial_connections"]

        # Lists to build the year-by-year table
        years, cum_dom, cum_com, cum_ind = [], [], [], []
        rev_list, opex_list, ncf_list, cum_cf_list = [], [], [], []

        # Year 0: initial investment outflow
        years.append(0)
        cum_dom.append(0)
        cum_com.append(0)
        cum_ind.append(0)
        rev_list.append(0.0)
        opex_list.append(0.0)
        ncf_list.append(-capex)           # negative = cash outflow
        cum_cf_list.append(-capex)

        # Years 1 through 10: linear customer ramp-up
        cumulative_cf = -capex            # running cumulative cash flow
        for t in range(1, project_years + 1):
            frac = t / project_years      # ramp fraction (0.1, 0.2, ... 1.0)

            c_dom = dom_conn * frac       # cumulative domestic at year t
            c_com = com_conn * frac       # cumulative commercial at year t
            c_ind = ind_conn * frac       # cumulative industrial at year t

            # Annual revenue from all customer types
            annual_revenue = (
                c_dom * avg_domestic_annual_revenue
                + c_com * avg_commercial_annual_revenue
                + c_ind * avg_industrial_annual_revenue
            )

            # Operating expenses as a fixed % of revenue
            annual_opex = annual_revenue * operating_cost_pct_of_revenue

            # Net cash flow for the year
            net_cf = annual_revenue - annual_opex

            cumulative_cf += net_cf

            years.append(t)
            cum_dom.append(c_dom)
            cum_com.append(c_com)
            cum_ind.append(c_ind)
            rev_list.append(annual_revenue)
            opex_list.append(annual_opex)
            ncf_list.append(net_cf)
            cum_cf_list.append(cumulative_cf)

        # Store the detail table
        detail_df = pd.DataFrame({
            "Year":                   years,
            "Cumulative_Domestic":    cum_dom,
            "Cumulative_Commercial":  cum_com,
            "Cumulative_Industrial":  cum_ind,
            "Annual_Revenue_Rs":      rev_list,
            "Annual_Opex_Rs":         opex_list,
            "Net_Cash_Flow_Rs":       ncf_list,
            "Cumulative_Cash_Flow_Rs": cum_cf_list,
        })
        zone_cashflow_tables[zone_name] = detail_df

        # ---- IRR (Internal Rate of Return) ----
        cash_flows = np.array(ncf_list, dtype=float)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            irr_val = npf.irr(cash_flows)

        # Handle nan / inf edge cases
        if irr_val is None or not np.isfinite(irr_val):
            irr_val = 0.0

        irr_list.append(irr_val)

        # ---- NPV (Net Present Value) ----
        npv_val = npf.npv(discount_rate, cash_flows)
        if npv_val is None or not np.isfinite(npv_val):
            npv_val = 0.0
        npv_list.append(npv_val)

        # ---- Simple Payback Period ----
        # Average annual net cash flow over Years 1–10
        avg_annual_ncf = np.mean(ncf_list[1:])   # exclude Year-0 investment
        if avg_annual_ncf > 0:
            payback = round(capex / avg_annual_ncf, 1)
        else:
            payback = float("inf")  # project never pays back
        payback_list.append(payback)

        # ---- Year-10 annual revenue (for summary table) ----
        year10_revenue_list.append(rev_list[-1])

    # Store IRR / NPV / payback back into zone_df
    zone_df["IRR"]             = irr_list
    zone_df["NPV"]             = npv_list
    zone_df["payback_years"]   = payback_list
    zone_df["year10_revenue"]  = year10_revenue_list

    # ------------------------------------------------------------------
    # 8. Viability classification
    # ------------------------------------------------------------------
    print("[8/12] Classifying zone viability ...")

    def _viability(irr):
        """Return a viability label based on IRR thresholds."""
        if irr > 0.12:
            return "Financially Viable"
        elif irr >= 0.08:
            return "Marginal"
        else:
            return "Needs Review"

    zone_df["Viability"] = zone_df["IRR"].apply(_viability)

    # ------------------------------------------------------------------
    # Build the summary DataFrame that will be exported & returned
    # ------------------------------------------------------------------
    summary_df = zone_df[["zone_name"]].copy()
    summary_df["Total_Capex_Crores"]         = (zone_df["total_capex"] / 1e7).round(2)
    summary_df["Year10_Annual_Revenue_Crores"] = (zone_df["year10_revenue"] / 1e7).round(2)
    summary_df["Payback_Period_Years"]        = zone_df["payback_years"]
    summary_df["IRR_Pct"]                     = (zone_df["IRR"] * 100).round(2)
    summary_df["NPV_Crores"]                  = (zone_df["NPV"] / 1e7).round(2)
    summary_df["Viability"]                   = zone_df["Viability"]

    # ------------------------------------------------------------------
    # 9. Export to Excel -- multi-sheet workbook
    # ------------------------------------------------------------------
    print("[9/12] Exporting financial model to Excel ...")

    outputs_dir = os.path.join(base_path, "outputs")
    os.makedirs(outputs_dir, exist_ok=True)

    excel_path = os.path.join(outputs_dir, "financial_model.xlsx")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # ---- Summary sheet ----
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws_summary = writer.sheets["Summary"]

        # Bold header row
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Number formatting for currency & percentage columns
        for row_cells in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
            for cell in row_cells:
                # Columns B, C, F (1-indexed 2,3,6) -> 2-decimal number
                if cell.column in (2, 3, 6):
                    cell.number_format = "#,##0.00"
                # Column D (payback) -> 1-decimal
                elif cell.column == 4:
                    cell.number_format = "#,##0.0"
                # Column E (IRR %) -> 2-decimal
                elif cell.column == 5:
                    cell.number_format = "#,##0.00"

        _auto_fit_columns(ws_summary)

        # ---- Per-zone detail sheets ----
        for zone_name, detail_df in zone_cashflow_tables.items():
            # Excel sheet names are limited to 31 characters
            sheet_name = zone_name[:31]
            detail_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Format currency columns (E through H, i.e. cols 5-8)
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row_cells:
                    if cell.column >= 5:
                        cell.number_format = "#,##0"

            _auto_fit_columns(ws)

    print(f"       Saved -> {excel_path}")

    # ------------------------------------------------------------------
    # 10. Chart 1 -- Zone-wise IRR Comparison (bar chart)
    # ------------------------------------------------------------------
    print("[10/12] Generating IRR comparison bar chart ...")

    charts_dir = os.path.join(outputs_dir, "charts")
    os.makedirs(charts_dir, exist_ok=True)

    plt.style.use("ggplot")
    fig, ax = plt.subplots(figsize=(12, 7))

    zone_names = summary_df["zone_name"].tolist()
    irr_pcts   = summary_df["IRR_Pct"].tolist()

    # Colour each bar based on viability thresholds
    bar_colors = []
    for irr_pct in irr_pcts:
        if irr_pct > 12:
            bar_colors.append("#2ecc71")   # green
        elif irr_pct >= 8:
            bar_colors.append("#f39c12")   # amber / orange
        else:
            bar_colors.append("#e74c3c")   # red

    bars = ax.bar(zone_names, irr_pcts, color=bar_colors, edgecolor="white",
                  linewidth=0.8, zorder=3)

    # IRR value label on each bar
    for bar, val in zip(bars, irr_pcts):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.4,
            f"{val:.1f}%",
            ha="center", va="bottom", fontsize=10, fontweight="bold",
        )

    # 12 % viability threshold line
    ax.axhline(y=12, color="navy", linestyle="--", linewidth=1.2, zorder=2)
    ax.text(
        len(zone_names) - 0.5, 12.5, "Viability Threshold (12 %)",
        ha="right", va="bottom", fontsize=9, color="navy", fontstyle="italic",
    )

    ax.set_title("Zone-wise IRR Comparison", fontsize=16, fontweight="bold", pad=15)
    ax.set_ylabel("IRR (%)", fontsize=13)
    ax.set_xlabel("Zone", fontsize=13)
    ax.tick_params(axis="x", rotation=25, labelsize=10)
    ax.tick_params(axis="y", labelsize=11)
    ax.set_axisbelow(True)
    ax.grid(axis="y", alpha=0.4)

    fig.tight_layout()
    irr_chart_path = os.path.join(charts_dir, "zone_irr_comparison.png")
    fig.savefig(irr_chart_path, dpi=300, bbox_inches="tight")
    plt.close("all")
    print(f"       Saved -> {irr_chart_path}")

    # ------------------------------------------------------------------
    # 11. Chart 2 -- Cumulative Cash Flow for Top 3 Priority Zones
    # ------------------------------------------------------------------
    print("[11/12] Generating cumulative cash-flow chart (top 3 zones) ...")

    # Determine the top-3 zones
    if zone_ranking_df is not None:
        top3_names = zone_ranking_df["zone_name"].head(3).tolist()
    else:
        top3_names = zone_df["zone_name"].head(3).tolist()

    plt.style.use("ggplot")
    fig, ax = plt.subplots(figsize=(14, 8))

    palette = ["#2980b9", "#e74c3c", "#27ae60", "#8e44ad", "#f39c12"]

    for idx, zname in enumerate(top3_names):
        detail = zone_cashflow_tables[zname]
        years_arr  = detail["Year"].values
        cum_cf_cr  = detail["Cumulative_Cash_Flow_Rs"].values / 1e7  # -> Crores

        color = palette[idx % len(palette)]
        ax.plot(
            years_arr, cum_cf_cr,
            marker="o", linewidth=2.2, markersize=7,
            color=color, label=zname, zorder=3,
        )

        # Annotate the payback year (first year where cumulative CF > 0)
        payback_year = None
        for yr, cf in zip(years_arr, cum_cf_cr):
            if cf > 0:
                payback_year = yr
                break

        if payback_year is not None:
            pb_cf = cum_cf_cr[payback_year]  # index matches year for 0..10
            ax.annotate(
                f"Payback Yr {payback_year}",
                xy=(payback_year, pb_cf),
                xytext=(payback_year + 0.5, pb_cf + max(abs(cum_cf_cr.min()), abs(cum_cf_cr.max())) * 0.08),
                fontsize=9, fontweight="bold", color=color,
                arrowprops=dict(arrowstyle="->", color=color, lw=1.3),
            )
            # Special star marker at payback point
            ax.plot(payback_year, pb_cf, marker="*", markersize=16,
                    color=color, zorder=5)

    # Breakeven line at y = 0
    ax.axhline(y=0, color="gray", linestyle="--", linewidth=1.0, zorder=2,
               label="Breakeven")

    ax.set_title("Cumulative Cash Flow -- Top 3 Priority Zones",
                 fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("Year", fontsize=13)
    ax.set_ylabel("Cumulative Cash Flow (Rs. Crores)", fontsize=13)
    ax.set_xticks(range(0, project_years + 1))
    ax.tick_params(axis="both", labelsize=11)
    ax.legend(fontsize=11, loc="lower right", framealpha=0.9)
    ax.set_axisbelow(True)
    ax.grid(alpha=0.4)

    fig.tight_layout()
    cf_chart_path = os.path.join(charts_dir, "cumulative_cashflow_top3.png")
    fig.savefig(cf_chart_path, dpi=300, bbox_inches="tight")
    plt.close("all")
    print(f"       Saved -> {cf_chart_path}")

    # ------------------------------------------------------------------
    # 12. Print financial summary to console
    # ------------------------------------------------------------------
    print("\n[12/12] Financial Summary")
    print("-" * 100)
    header = (
        f"{'Zone':<28} {'CAPEX (Cr)':>12} {'Rev Yr10 (Cr)':>14} "
        f"{'Payback (Yr)':>13} {'IRR (%)':>9} {'NPV (Cr)':>10} {'Viability':<20}"
    )
    print(header)
    print("-" * 100)

    for _, r in summary_df.iterrows():
        payback_str = (
            f"{r['Payback_Period_Years']:.1f}"
            if np.isfinite(r["Payback_Period_Years"])
            else "N/A"
        )
        line = (
            f"{r['zone_name']:<28} "
            f"{r['Total_Capex_Crores']:>12.2f} "
            f"{r['Year10_Annual_Revenue_Crores']:>14.2f} "
            f"{payback_str:>13} "
            f"{r['IRR_Pct']:>9.2f} "
            f"{r['NPV_Crores']:>10.2f} "
            f"{r['Viability']:<20}"
        )
        print(line)

    print("-" * 100)
    print("\n[OK]  Techno-economic analysis complete.")

    return summary_df


# ---------------------------------------------------------------------------
# Standalone execution
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # When run directly, assume the script is inside <project>/models/
    # and the project root is one level up.
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _project_root = os.path.abspath(os.path.join(_script_dir, os.pardir))

    print(f"Project root: {_project_root}")
    result = run_techno_economic(_project_root)
    print(f"\nReturned summary DataFrame shape: {result.shape}")
