"""
zone_prioritization.py
======================
Zone Prioritization Module for CGD (City Gas Distribution) Network Expansion.

This module evaluates geographic zones for pipeline expansion by computing a
weighted composite score from four factors:

    1. Household Density        (weight 0.30)
    2. Industrial Potential     (weight 0.30)
    3. Existing Penetration     (weight 0.20)
    4. Proximity to Network     (weight 0.20)

Each raw factor is min-max normalised to a 0-100 scale, multiplied by its
weight, and summed into a Final_Score.  Zones are ranked, assigned an
expansion-year recommendation, and the results are exported to:

    * outputs/zone_ranking.xlsx           – formatted workbook
    * outputs/charts/zone_priority_ranking.png  – horizontal bar chart
    * outputs/charts/zone_score_breakdown.png   – grouped bar chart

Usage
-----
    from models.zone_prioritization import run_zone_prioritization
    df = run_zone_prioritization(base_path=r"C:\\project_root")

Author : Auto-generated
Date   : 2026-06-27
"""

import os
import sys

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers


# ---------------------------------------------------------------------------
# Helper: min-max normalisation to 0-100
# ---------------------------------------------------------------------------
def _min_max_normalise(series: pd.Series) -> pd.Series:
    """
    Apply min-max normalisation to a pandas Series.

    Formula:  score = (value - min) / (max - min) * 100

    If all values are identical (max == min), every element is mapped to 0.0
    to avoid division-by-zero.

    Parameters
    ----------
    series : pd.Series
        Raw numeric values to normalise.

    Returns
    -------
    pd.Series
        Normalised values in [0, 100].
    """
    min_val = series.min()
    max_val = series.max()
    if max_val == min_val:
        # All values are the same – return zeros to avoid division by zero
        return pd.Series(0.0, index=series.index)
    return ((series - min_val) / (max_val - min_val)) * 100


# ---------------------------------------------------------------------------
# Helper: assign expansion-year recommendation based on priority rank
# ---------------------------------------------------------------------------
def _assign_recommendation(rank: int) -> str:
    """
    Map a 1-based Priority_Rank to an expansion-year recommendation string.

    Parameters
    ----------
    rank : int
        Priority rank (1 = highest priority).

    Returns
    -------
    str
        One of 'Year 1 Expansion', 'Year 2 Expansion', 'Year 3 Expansion'.
    """
    if rank <= 3:
        return "Year 1 Expansion"
    elif rank <= 5:
        return "Year 2 Expansion"
    else:
        return "Year 3 Expansion"


# ---------------------------------------------------------------------------
# Chart 1: Horizontal bar – Zone Priority Ranking
# ---------------------------------------------------------------------------
def _create_priority_bar_chart(df: pd.DataFrame, output_path: str) -> None:
    """
    Create a horizontal bar chart showing zone priority scores.

    Bars are coloured by recommendation year:
        * Green  -> Year 1
        * Amber  -> Year 2
        * Red    -> Year 3

    The chart is saved to *output_path* at 300 dpi.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain columns 'zone_name', 'Final_Score', 'Recommendation'.
    output_path : str
        Absolute path for the saved PNG image.
    """
    print("  [Chart 1] Generating zone priority ranking bar chart ...")

    # ---- colour map keyed by recommendation string ----
    colour_map = {
        "Year 1 Expansion": "#2ecc71",   # green
        "Year 2 Expansion": "#f39c12",   # amber / orange
        "Year 3 Expansion": "#e74c3c",   # red
    }

    # Sort ascending so that the highest-score zone appears at the TOP of the
    # horizontal bar chart (matplotlib draws bars bottom-to-top).
    plot_df = df.sort_values("Final_Score", ascending=True).reset_index(drop=True)

    # Build a colour list aligned with the (ascending-sorted) DataFrame
    bar_colours = [colour_map.get(rec, "#95a5a6") for rec in plot_df["Recommendation"]]

    # ---- draw ----
    plt.style.use("ggplot")
    fig, ax = plt.subplots(figsize=(12, 8))

    y_positions = np.arange(len(plot_df))
    bars = ax.barh(
        y_positions,
        plot_df["Final_Score"],
        color=bar_colours,
        edgecolor="white",
        height=0.6,
    )

    # Zone names on the Y axis
    ax.set_yticks(y_positions)
    ax.set_yticklabels(plot_df["zone_name"], fontsize=11)

    # Axis labels and title
    ax.set_xlabel("Final Priority Score", fontsize=12, fontweight="bold")
    ax.set_title("Zone Expansion Priority Ranking", fontsize=16, fontweight="bold", pad=15)

    # Annotate score values at the end of each bar
    for bar_obj in bars:
        width = bar_obj.get_width()
        ax.text(
            width + 0.5,
            bar_obj.get_y() + bar_obj.get_height() / 2,
            f"{width:.2f}",
            va="center",
            ha="left",
            fontsize=10,
            fontweight="bold",
        )

    # Add a small right-margin so annotations are not clipped
    ax.set_xlim(0, plot_df["Final_Score"].max() * 1.12)

    plt.tight_layout()
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close("all")
    print(f"  [Chart 1] Saved -> {output_path}")


# ---------------------------------------------------------------------------
# Chart 2: Grouped bar – Score Breakdown by Factor
# ---------------------------------------------------------------------------
def _create_score_breakdown_chart(df: pd.DataFrame, output_path: str) -> None:
    """
    Create a grouped bar chart showing each factor's WEIGHTED contribution
    to the Final_Score for every zone.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain the four weighted score columns and 'zone_name'.
    output_path : str
        Absolute path for the saved PNG image.
    """
    print("  [Chart 2] Generating score breakdown grouped bar chart ...")

    # ---- factor definitions (column name -> legend label, colour) ----
    factors = [
        ("Weighted_Household_Density", "Household Density (30%)", "#3498db"),
        ("Weighted_Industrial",        "Industrial Potential (30%)", "#e67e22"),
        ("Weighted_Penetration",       "Existing Penetration (20%)", "#2ecc71"),
        ("Weighted_Proximity",         "Proximity (20%)", "#9b59b6"),
    ]
    n_zones = len(df)
    n_factors = len(factors)
    bar_width = 0.18  # width of each individual bar
    x = np.arange(n_zones)

    plt.style.use("ggplot")
    fig, ax = plt.subplots(figsize=(14, 9))

    # Draw one set of bars per factor, offset horizontally
    for i, (col, label, colour) in enumerate(factors):
        offset = (i - n_factors / 2 + 0.5) * bar_width
        ax.bar(
            x + offset,
            df[col],
            width=bar_width,
            label=label,
            color=colour,
            edgecolor="white",
        )

    # X-axis: zone names
    ax.set_xticks(x)
    ax.set_xticklabels(df["zone_name"], fontsize=10, rotation=30, ha="right")

    # Axis labels and title
    ax.set_ylabel("Score Contribution (Weighted)", fontsize=12, fontweight="bold")
    ax.set_xlabel("Zone", fontsize=12, fontweight="bold")
    ax.set_title(
        "Zone Score Breakdown by Factor",
        fontsize=16,
        fontweight="bold",
        pad=15,
    )

    ax.legend(fontsize=10, loc="upper right", framealpha=0.9)

    plt.tight_layout()
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close("all")
    print(f"  [Chart 2] Saved -> {output_path}")


# ---------------------------------------------------------------------------
# Excel export with openpyxl formatting
# ---------------------------------------------------------------------------
def _export_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Write the ranking DataFrame to an Excel workbook with professional
    formatting via openpyxl:

        * Bold headers
        * Auto-adjusted column widths
        * Numeric formatting (2 decimal places)
        * Colour-coded Recommendation column (green / amber / red)

    Parameters
    ----------
    df : pd.DataFrame
        The fully scored and ranked DataFrame.
    output_path : str
        Absolute path for the .xlsx file.
    """
    print("  [Excel] Exporting formatted workbook ...")

    # --- Step 1: write raw data with pandas ----------------------------------
    df.to_excel(output_path, index=False, sheet_name="Zone Ranking", engine="openpyxl")

    # --- Step 2: open with openpyxl and apply formatting ---------------------
    wb = load_workbook(output_path)
    ws = wb.active

    # -- header styling -------------------------------------------------------
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # -- auto-adjust column widths --------------------------------------------
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        # Add a little padding
        ws.column_dimensions[col_letter].width = max_length + 4

    # -- number formatting for numeric cells ----------------------------------
    # Identify columns that should receive number formatting (all score cols)
    score_columns = {
        "Household_Density_Score",
        "Industrial_Score",
        "Penetration_Score",
        "Proximity_Score",
        "Weighted_Household_Density",
        "Weighted_Industrial",
        "Weighted_Penetration",
        "Weighted_Proximity",
        "Final_Score",
        "household_density",
        "industrial_equivalent",
        "proximity_raw",
    }

    # Build a mapping from column index (1-based) to column name
    col_name_map = {}
    for idx, cell in enumerate(ws[1], start=1):
        col_name_map[idx] = cell.value

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name = col_name_map.get(cell.column)
            if col_name in score_columns:
                cell.number_format = "0.00"

    # -- colour Recommendation column -----------------------------------------
    # Locate the Recommendation column index (1-based)
    rec_col_idx = None
    for idx, name in col_name_map.items():
        if name == "Recommendation":
            rec_col_idx = idx
            break

    if rec_col_idx is not None:
        # Define fills
        green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        amber_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        white_font = Font(bold=True, color="FFFFFF", size=11)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=rec_col_idx, max_col=rec_col_idx):
            for cell in row:
                if cell.value == "Year 1 Expansion":
                    cell.fill = green_fill
                    cell.font = white_font
                elif cell.value == "Year 2 Expansion":
                    cell.fill = amber_fill
                    cell.font = white_font
                elif cell.value == "Year 3 Expansion":
                    cell.fill = red_fill
                    cell.font = white_font

    # --- Step 3: save --------------------------------------------------------
    wb.save(output_path)
    print(f"  [Excel] Saved -> {output_path}")


# ===========================================================================
# MAIN FUNCTION
# ===========================================================================
def run_zone_prioritization(base_path: str) -> pd.DataFrame:
    """
    Execute the full zone prioritization pipeline.

    Steps
    -----
    1. Read ``inputs/zone_data.csv``
    2. Compute four factor scores (min-max normalised 0-100)
    3. Calculate weighted composite Final_Score
    4. Rank zones and assign expansion-year recommendations
    5. Export formatted Excel workbook
    6. Generate two charts (priority ranking + score breakdown)
    7. Print ranking table to console
    8. Return the sorted DataFrame

    Parameters
    ----------
    base_path : str
        Project root directory.  All relative paths (inputs/, outputs/) are
        resolved from here using ``os.path.join``.

    Returns
    -------
    pd.DataFrame
        The scored and ranked zone DataFrame, sorted by Final_Score descending.
    """
    print("=" * 70)
    print(" CGD ZONE PRIORITIZATION MODULE")
    print("=" * 70)

    # ------------------------------------------------------------------
    # 1. READ INPUT DATA
    # ------------------------------------------------------------------
    input_path = os.path.join(base_path, "inputs", "zone_data.csv")
    print(f"\n[Step 1] Reading zone data from: {input_path}")

    df = pd.read_csv(input_path)
    print(f"  -> Loaded {len(df)} zones with columns: {list(df.columns)}")

    # ------------------------------------------------------------------
    # 2. CALCULATE RAW FACTOR VALUES
    # ------------------------------------------------------------------
    print("\n[Step 2] Computing raw factor values ...")

    # Factor 1 -- Household Density (households per sq km)
    df["household_density"] = df["total_households"] / df["area_sqkm"]
    print(f"  -> household_density range: "
          f"{df['household_density'].min():.2f} – {df['household_density'].max():.2f}")

    # Factor 2 -- Industrial Equivalent
    # Converts industrial units into an equivalent household count to allow
    # comparison on the same scale: each industrial unit ~ 500 households.
    df["industrial_equivalent"] = (df["industrial_units"] * 500) + df["total_households"]
    print(f"  -> industrial_equivalent range: "
          f"{df['industrial_equivalent'].min():.0f} – {df['industrial_equivalent'].max():.0f}")

    # Factor 3 -- Existing Penetration (used directly from the CSV column)
    # No derived column needed; we normalise existing_penetration_pct below.

    # Factor 4 -- Proximity (inverse of distance -> closer = higher score)
    df["proximity_raw"] = 1.0 / df["distance_from_network_km"]
    print(f"  -> proximity_raw range: "
          f"{df['proximity_raw'].min():.4f} – {df['proximity_raw'].max():.4f}")

    # ------------------------------------------------------------------
    # 3. MIN-MAX NORMALISE EACH FACTOR TO 0-100
    # ------------------------------------------------------------------
    print("\n[Step 3] Normalising factor scores (0-100) ...")

    df["Household_Density_Score"] = _min_max_normalise(df["household_density"])
    df["Industrial_Score"] = _min_max_normalise(df["industrial_equivalent"])
    df["Penetration_Score"] = _min_max_normalise(df["existing_penetration_pct"])
    df["Proximity_Score"] = _min_max_normalise(df["proximity_raw"])

    # ------------------------------------------------------------------
    # 4. COMPUTE WEIGHTED CONTRIBUTIONS AND FINAL SCORE
    # ------------------------------------------------------------------
    print("\n[Step 4] Computing weighted scores and Final_Score ...")

    # Weighted contributions (stored for the breakdown chart)
    df["Weighted_Household_Density"] = df["Household_Density_Score"] * 0.30
    df["Weighted_Industrial"] = df["Industrial_Score"] * 0.30
    df["Weighted_Penetration"] = df["Penetration_Score"] * 0.20
    df["Weighted_Proximity"] = df["Proximity_Score"] * 0.20

    # Composite final score
    df["Final_Score"] = (
        df["Weighted_Household_Density"]
        + df["Weighted_Industrial"]
        + df["Weighted_Penetration"]
        + df["Weighted_Proximity"]
    )

    # ------------------------------------------------------------------
    # 5. ROUND ALL SCORES TO 2 DECIMAL PLACES
    # ------------------------------------------------------------------
    print("\n[Step 5] Rounding scores to 2 decimal places ...")

    score_cols = [
        "household_density",
        "industrial_equivalent",
        "proximity_raw",
        "Household_Density_Score",
        "Industrial_Score",
        "Penetration_Score",
        "Proximity_Score",
        "Weighted_Household_Density",
        "Weighted_Industrial",
        "Weighted_Penetration",
        "Weighted_Proximity",
        "Final_Score",
    ]
    df[score_cols] = df[score_cols].round(2)

    # ------------------------------------------------------------------
    # 6. SORT BY FINAL_SCORE DESCENDING AND ASSIGN RANK
    # ------------------------------------------------------------------
    print("\n[Step 6] Sorting zones by Final_Score (descending) ...")

    df = df.sort_values("Final_Score", ascending=False).reset_index(drop=True)
    df["Priority_Rank"] = range(1, len(df) + 1)

    # ------------------------------------------------------------------
    # 7. ASSIGN EXPANSION-YEAR RECOMMENDATION
    # ------------------------------------------------------------------
    print("\n[Step 7] Assigning expansion-year recommendations ...")

    df["Recommendation"] = df["Priority_Rank"].apply(_assign_recommendation)

    for _, row in df.iterrows():
        print(f"  Rank {row['Priority_Rank']}: {row['zone_name']} -> {row['Recommendation']}")

    # ------------------------------------------------------------------
    # 8. ENSURE OUTPUT DIRECTORIES EXIST
    # ------------------------------------------------------------------
    outputs_dir = os.path.join(base_path, "outputs")
    charts_dir = os.path.join(outputs_dir, "charts")
    os.makedirs(outputs_dir, exist_ok=True)
    os.makedirs(charts_dir, exist_ok=True)
    print(f"\n[Step 8] Output directories ready: {outputs_dir}")

    # ------------------------------------------------------------------
    # 9. EXPORT FORMATTED EXCEL WORKBOOK
    # ------------------------------------------------------------------
    excel_path = os.path.join(outputs_dir, "zone_ranking.xlsx")
    print(f"\n[Step 9] Exporting Excel workbook ...")
    _export_to_excel(df, excel_path)

    # ------------------------------------------------------------------
    # 10. GENERATE CHARTS
    # ------------------------------------------------------------------
    print("\n[Step 10] Generating charts ...")

    chart1_path = os.path.join(charts_dir, "zone_priority_ranking.png")
    _create_priority_bar_chart(df, chart1_path)

    chart2_path = os.path.join(charts_dir, "zone_score_breakdown.png")
    _create_score_breakdown_chart(df, chart2_path)

    # ------------------------------------------------------------------
    # 11. PRINT RANKING TABLE TO CONSOLE
    # ------------------------------------------------------------------
    print("\n[Step 11] Zone Ranking Summary")
    print("=" * 70)

    # Select the most informative columns for console display
    display_cols = [
        "Priority_Rank",
        "zone_name",
        "Household_Density_Score",
        "Industrial_Score",
        "Penetration_Score",
        "Proximity_Score",
        "Final_Score",
        "Recommendation",
    ]
    print(df[display_cols].to_string(index=False))

    print("\n" + "=" * 70)
    print(" ZONE PRIORITIZATION COMPLETE")
    print("=" * 70)
    print(f"  Excel  : {excel_path}")
    print(f"  Chart 1: {chart1_path}")
    print(f"  Chart 2: {chart2_path}")
    print("=" * 70)

    # ------------------------------------------------------------------
    # 12. RETURN THE SORTED DATAFRAME
    # ------------------------------------------------------------------
    return df


# ===========================================================================
# CLI entry point
# ===========================================================================
if __name__ == "__main__":
    # When run directly, assume the project root is two levels up from this
    # file's location (models/ -> cgd_project/).  The user may also pass an
    # explicit path as the first CLI argument.
    if len(sys.argv) > 1:
        project_root = sys.argv[1]
    else:
        # Derive from file location: .../models/zone_prioritization.py -> .../
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    print(f"Project root: {project_root}\n")
    result_df = run_zone_prioritization(project_root)
    print(f"\nReturned DataFrame shape: {result_df.shape}")
